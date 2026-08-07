#!/usr/bin/env python3
"""Generate the integrated, paper-ready analysis of official Dhaka DoE data."""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any

os.environ.setdefault("MPLCONFIGDIR", "/tmp/dhaka-aqi-matplotlib")

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scipy.stats as stats
import statsmodels.api as sm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from statsmodels.stats.proportion import proportion_confint

ROOT = Path(__file__).resolve().parents[1]
PROCESSED = ROOT / "data/processed"
CONTEXT = ROOT / "data/context"
OUTPUT = ROOT / "analysis"
FIGURES = OUTPUT / "figures"

SEASONS = ["Winter", "Pre-monsoon", "Monsoon", "Post-monsoon"]
CORE_VARIABLES = {
    "aqi": ("aqi_mean", r"$\mathrm{AQI}$", "index"),
    "pm25": ("pm25_mean", r"$\mathrm{PM}_{2.5}$", r"$\mu\mathrm{g}\,\mathrm{m}^{-3}$"),
    "pm10": ("pm10_mean", r"$\mathrm{PM}_{10}$", r"$\mu\mathrm{g}\,\mathrm{m}^{-3}$"),
}
PLAIN_LABELS = {"aqi": "AQI", "pm25": "PM2.5", "pm10": "PM10"}
PLAIN_UNITS = {"aqi": "index", "pm25": "µg/m³", "pm10": "µg/m³"}
COLORS = {"aqi": "#5E3C99", "pm25": "#D73027", "pm10": "#F46D43"}
HISTORICAL_POLLUTANTS = {
    "pm25": ("pm25_mean", "PM2.5", "µg/m³"),
    "pm10": ("pm10_mean", "PM10", "µg/m³"),
    "no2": ("no2_mean", "NO2", "ppb"),
    "so2": ("so2_mean", "SO2", "ppb"),
    "co": ("co_mean", "CO", "ppm"),
    "o3": ("o3_mean", "O3", "ppb"),
}
PARTICULATE_ERAS = {
    "Historical public-report era": (2013, 2019),
    "Recent joint-measurement era": (2022, 2025),
}


def season_for_month(month: int) -> str:
    if month in {12, 1, 2}:
        return "Winter"
    if month in {3, 4, 5}:
        return "Pre-monsoon"
    if month in {6, 7, 8, 9}:
        return "Monsoon"
    return "Post-monsoon"


def fdr(p_values: pd.Series) -> pd.Series:
    result = pd.Series(np.nan, index=p_values.index, dtype=float)
    ordered = p_values.dropna().sort_values()
    if ordered.empty:
        return result
    adjusted = ordered * len(ordered) / np.arange(1, len(ordered) + 1)
    result.loc[ordered.index] = adjusted.iloc[::-1].cummin().iloc[::-1].clip(upper=1)
    return result


def bootstrap_mean_ci(values: pd.Series, seed: int, draws: int = 5_000) -> tuple[float, float]:
    array = values.dropna().to_numpy(dtype=float)
    if len(array) < 2:
        return np.nan, np.nan
    rng = np.random.default_rng(seed)
    means = rng.choice(array, size=(draws, len(array)), replace=True).mean(axis=1)
    return tuple(np.quantile(means, [0.025, 0.975]))


def descriptive_table(wide: pd.DataFrame) -> pd.DataFrame:
    recent = wide[wide["year"].between(2022, 2025)]
    rows: list[dict[str, Any]] = []
    for key, (column, _, _) in CORE_VARIABLES.items():
        values = recent[column].dropna()
        rows.append(
            {
                "variable": key,
                "label": PLAIN_LABELS[key],
                "unit": PLAIN_UNITS[key],
                "analysis_period": "2022-01 through 2025-12",
                "n_months": len(values),
                "mean": values.mean(),
                "standard_deviation": values.std(ddof=1),
                "median": values.median(),
                "q1": values.quantile(0.25),
                "q3": values.quantile(0.75),
                "minimum": values.min(),
                "maximum": values.max(),
            }
        )
    return pd.DataFrame(rows)


def seasonal_tables(wide: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    recent = wide[wide["year"].between(2022, 2025)].copy()
    recent["season"] = pd.Categorical(recent["season"], SEASONS, ordered=True)
    summaries: list[dict[str, Any]] = []
    tests: list[dict[str, Any]] = []
    for variable_index, (key, (column, _, _)) in enumerate(CORE_VARIABLES.items()):
        groups = []
        for season_index, season in enumerate(SEASONS):
            values = recent.loc[recent["season"].eq(season), column].dropna()
            groups.append(values.to_numpy())
            low, high = bootstrap_mean_ci(values, 20260806 + 10 * variable_index + season_index)
            summaries.append(
                {
                    "variable": key,
                    "label": PLAIN_LABELS[key],
                    "unit": PLAIN_UNITS[key],
                    "season": season,
                    "n_months": len(values),
                    "mean": values.mean(),
                    "mean_bootstrap_ci_low": low,
                    "mean_bootstrap_ci_high": high,
                    "median": values.median(),
                    "q1": values.quantile(0.25),
                    "q3": values.quantile(0.75),
                }
            )
        statistic, p_value = stats.kruskal(*groups)
        tests.append(
            {
                "variable": key,
                "test": "Kruskal-Wallis across four seasons",
                "statistic": statistic,
                "p_value": p_value,
                "n_months": sum(map(len, groups)),
            }
        )
    tests_frame = pd.DataFrame(tests)
    tests_frame["q_value_bh"] = fdr(tests_frame["p_value"])
    return pd.DataFrame(summaries), tests_frame


def long_term_particulate_tables(wide: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Summarize the two defensible particulate eras without bridging the archive gap."""
    summary_rows: list[dict[str, Any]] = []
    annual_rows: list[dict[str, Any]] = []
    seasonal_rows: list[dict[str, Any]] = []

    for era, (start_year, end_year) in PARTICULATE_ERAS.items():
        period = wide[wide["year"].between(start_year, end_year)].copy()
        for key in ("pm25", "pm10"):
            column = CORE_VARIABLES[key][0]
            values = period[column].dropna()
            summary_rows.append(
                {
                    "era": era,
                    "period": f"{start_year}-{end_year}",
                    "variable": key,
                    "label": PLAIN_LABELS[key],
                    "unit": PLAIN_UNITS[key],
                    "n_months": len(values),
                    "mean": values.mean(),
                    "standard_deviation": values.std(ddof=1),
                    "median": values.median(),
                    "q1": values.quantile(0.25),
                    "q3": values.quantile(0.75),
                    "minimum": values.min(),
                    "maximum": values.max(),
                    "mean_station_count": period.loc[
                        period[column].notna(), f"{key}_station_count"
                    ].mean(),
                    "mean_data_capture_pct": period.loc[
                        period[column].notna(), f"{key}_mean_data_capture_pct"
                    ].mean(),
                }
            )

            for year, group in period.groupby("year"):
                annual_values = group[column].dropna()
                annual_rows.append(
                    {
                        "era": era,
                        "year": year,
                        "variable": key,
                        "unit": PLAIN_UNITS[key],
                        "n_months": len(annual_values),
                        "mean": annual_values.mean(),
                        "median": annual_values.median(),
                        "minimum": annual_values.min(),
                        "maximum": annual_values.max(),
                    }
                )

            for season in SEASONS:
                season_values = period.loc[period["season"].eq(season), column].dropna()
                seasonal_rows.append(
                    {
                        "era": era,
                        "period": f"{start_year}-{end_year}",
                        "variable": key,
                        "unit": PLAIN_UNITS[key],
                        "season": season,
                        "n_months": len(season_values),
                        "mean": season_values.mean(),
                        "median": season_values.median(),
                        "q1": season_values.quantile(0.25),
                        "q3": season_values.quantile(0.75),
                    }
                )

    return {
        "particulate_era_summary": pd.DataFrame(summary_rows),
        "particulate_annual": pd.DataFrame(annual_rows),
        "particulate_era_season": pd.DataFrame(seasonal_rows),
    }


def prepare_daily(
    daily: pd.DataFrame, monthly_report: pd.DataFrame, wide: pd.DataFrame
) -> pd.DataFrame:
    """Select the daily source used for each monthly AQI aggregate."""
    selected = daily[daily["selected_record"].astype(str).str.lower().eq("true")].copy()
    selected["aqi"] = pd.to_numeric(selected["aqi"], errors="coerce")
    selected["month_start"] = selected["report_date"].dt.to_period("M").dt.to_timestamp()
    selected["source_basis"] = "standalone_daily_archive_report_date"
    selected = selected[
        [
            "report_date",
            "aqi",
            "source_basis",
            "source_url",
            "source_sha256",
            "responsible_pollutant",
        ]
    ]

    table = monthly_report.copy()
    table["aqi"] = pd.to_numeric(table["aqi"], errors="coerce")
    table["report_date"] = table["aqi_date"]
    table["month_start"] = table["report_date"].dt.to_period("M").dt.to_timestamp()
    table["source_basis"] = "monthly_report_table_6"
    table["responsible_pollutant"] = pd.NA
    table = table[
        [
            "report_date",
            "aqi",
            "source_basis",
            "source_url",
            "source_sha256",
            "responsible_pollutant",
        ]
    ]

    candidates = pd.concat([table, selected], ignore_index=True)
    candidates["month_start"] = candidates["report_date"].dt.to_period("M").dt.to_timestamp()
    basis = wide.loc[wide["year"].between(2022, 2025), ["month_start", "aqi_source_basis"]]
    unified = candidates.merge(basis, on="month_start", how="inner")
    unified = unified[unified["source_basis"].eq(unified["aqi_source_basis"])]
    unified = unified.dropna(subset=["aqi"]).sort_values("report_date")
    unified = unified.drop_duplicates("report_date", keep="first")
    unified["year"] = unified["report_date"].dt.year
    unified["month"] = unified["report_date"].dt.month
    unified["season"] = unified["month"].map(season_for_month)
    return unified.reset_index(drop=True)


def daily_source_concordance(
    daily: pd.DataFrame, monthly_report: pd.DataFrame
) -> dict[str, pd.DataFrame]:
    selected = daily[daily["selected_record"].astype(str).str.lower().eq("true")].copy()
    selected["aqi"] = pd.to_numeric(selected["aqi"], errors="coerce")
    table = monthly_report.copy()
    table["aqi"] = pd.to_numeric(table["aqi"], errors="coerce")
    overlap = table[["aqi_date", "aqi"]].merge(
        selected[["report_date", "aqi", "source_url"]],
        left_on="aqi_date",
        right_on="report_date",
        suffixes=("_monthly_report", "_standalone"),
    )
    overlap = overlap.dropna(subset=["aqi_monthly_report", "aqi_standalone"])
    overlap["difference"] = overlap["aqi_monthly_report"] - overlap["aqi_standalone"]
    overlap["absolute_difference"] = overlap["difference"].abs()
    overlap["year"] = overlap["aqi_date"].dt.year

    rows = []
    analysis_overlap = overlap[overlap["year"].between(2022, 2025)]
    groups = [
        ("Analysis period, 2022-2025", analysis_overlap),
        ("All overlap through 2026", overlap),
        *overlap.groupby("year"),
    ]
    for label, group in groups:
        rows.append(
            {
                "period": str(label),
                "n_overlap_days": len(group),
                "exact_agreement_pct": 100 * group["difference"].eq(0).mean(),
                "mean_absolute_difference": group["absolute_difference"].mean(),
                "root_mean_squared_difference": np.sqrt(np.mean(group["difference"] ** 2)),
                "pearson_correlation": group[["aqi_monthly_report", "aqi_standalone"]]
                .corr()
                .iloc[0, 1],
                "maximum_absolute_difference": group["absolute_difference"].max(),
            }
        )
    discrepancies = overlap.nlargest(20, "absolute_difference")[
        [
            "aqi_date",
            "aqi_monthly_report",
            "aqi_standalone",
            "difference",
            "absolute_difference",
            "source_url",
        ]
    ]
    return {
        "daily_source_concordance": pd.DataFrame(rows),
        "daily_source_discrepancies": discrepancies,
    }


def standalone_daily_metadata(daily: pd.DataFrame) -> dict[str, pd.DataFrame]:
    selected = daily[daily["selected_record"].astype(str).str.lower().eq("true")].copy()
    selected = selected[selected["report_date"].dt.year.between(2023, 2025)]
    selected["year"] = selected["report_date"].dt.year
    selected["season"] = selected["report_date"].dt.month.map(season_for_month)
    selected["category_normalized"] = (
        selected["aqi_category_as_reported"].astype("string").str.upper().str.strip()
    )
    category = (
        selected.dropna(subset=["category_normalized"])
        .groupby(["year", "category_normalized"], as_index=False)
        .size()
        .rename(columns={"size": "reported_days"})
    )
    category["share_within_year_pct"] = (
        100 * category["reported_days"] / category.groupby("year")["reported_days"].transform("sum")
    )
    responsible = (
        selected.dropna(subset=["responsible_pollutant"])
        .groupby(["year", "season", "responsible_pollutant"], as_index=False)
        .size()
        .rename(columns={"size": "reported_days"})
    )
    responsible["share_within_year_season_pct"] = (
        100
        * responsible["reported_days"]
        / responsible.groupby(["year", "season"])["reported_days"].transform("sum")
    )
    return {
        "standalone_aqi_categories": category,
        "responsible_pollutant": responsible,
    }


def daily_burden_tables(daily: pd.DataFrame) -> dict[str, pd.DataFrame]:
    yearly_rows: list[dict[str, Any]] = []
    for year, group in daily.groupby("year"):
        expected = 366 if pd.Timestamp(year, 12, 31).is_leap_year else 365
        yearly_rows.append(
            {
                "year": year,
                "reported_days": len(group),
                "calendar_days": expected,
                "coverage_pct": 100 * len(group) / expected,
                "mean_aqi": group["aqi"].mean(),
                "median_aqi": group["aqi"].median(),
                "days_aqi_gt_100": int(group["aqi"].gt(100).sum()),
                "days_aqi_gt_150": int(group["aqi"].gt(150).sum()),
                "days_aqi_gt_200": int(group["aqi"].gt(200).sum()),
                "share_gt_150_pct": 100 * group["aqi"].gt(150).mean(),
                "maximum_aqi": group["aqi"].max(),
            }
        )

    monthly_rows: list[dict[str, Any]] = []
    for month, group in daily.groupby("month_start"):
        n = len(group)
        for threshold in (100, 150, 200):
            count = int(group["aqi"].gt(threshold).sum())
            low, high = proportion_confint(count, n, alpha=0.05, method="wilson")
            monthly_rows.append(
                {
                    "month_start": month,
                    "threshold": threshold,
                    "reported_days": n,
                    "days_above_threshold": count,
                    "share_pct": 100 * count / n,
                    "wilson_ci_low_pct": 100 * low,
                    "wilson_ci_high_pct": 100 * high,
                }
            )

    seasonal_rows: list[dict[str, Any]] = []
    for season in SEASONS:
        group = daily[daily["season"].eq(season)]
        seasonal_rows.append(
            {
                "season": season,
                "reported_days": len(group),
                "mean_aqi": group["aqi"].mean(),
                "median_aqi": group["aqi"].median(),
                "share_gt_100_pct": 100 * group["aqi"].gt(100).mean(),
                "share_gt_150_pct": 100 * group["aqi"].gt(150).mean(),
                "share_gt_200_pct": 100 * group["aqi"].gt(200).mean(),
            }
        )
    return {
        "daily_year": pd.DataFrame(yearly_rows),
        "monthly_exceedance": pd.DataFrame(monthly_rows),
        "daily_season": pd.DataFrame(seasonal_rows),
    }


def episode_table(daily: pd.DataFrame, threshold: float = 150) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    current: list[pd.Series] = []
    previous_date: pd.Timestamp | None = None
    previous_high = False
    for _, row in daily.iterrows():
        date = row["report_date"]
        high = row["aqi"] > threshold
        consecutive = previous_date is not None and date == previous_date + pd.Timedelta(days=1)
        if high and (not previous_high or not consecutive):
            if current:
                rows.append(_summarize_episode(current, threshold))
            current = [row]
        elif high:
            current.append(row)
        elif current:
            rows.append(_summarize_episode(current, threshold))
            current = []
        previous_date = date
        previous_high = high
    if current:
        rows.append(_summarize_episode(current, threshold))
    frame = pd.DataFrame(rows)
    if not frame.empty:
        frame.insert(0, "episode_id", np.arange(1, len(frame) + 1))
    return frame


def _summarize_episode(rows: list[pd.Series], threshold: float) -> dict[str, Any]:
    values = np.array([row["aqi"] for row in rows], dtype=float)
    start = rows[0]["report_date"]
    end = rows[-1]["report_date"]
    return {
        "start_date": start.date(),
        "end_date": end.date(),
        "season": season_for_month(start.month),
        "duration_days": len(rows),
        "mean_aqi": values.mean(),
        "peak_aqi": values.max(),
        "cumulative_aqi_above_150": np.maximum(values - threshold, 0).sum(),
        "definition": "consecutive reported calendar days with AQI > 150; missing days break episodes",
    }


def normalize_station(value: str) -> str:
    compact = "".join(character for character in str(value).lower() if character.isalnum())
    if "barc" in compact:
        return "BARC/Farmgate"
    if "sangshad" in compact:
        return "Sangshad Bhaban"
    if compact in {"doe", "departmentofenvironment"}:
        return "DoE"
    if "darus" in compact:
        return "Darus Salam"
    return str(value).strip()


def station_sensitivity_series(monthly: pd.DataFrame) -> pd.DataFrame:
    frame = monthly.copy()
    frame["report_month"] = pd.to_datetime(frame["report_month"])
    frame = frame[frame["parameter"].isin(["PM2.5", "PM10"])].copy()
    frame["station"] = frame["station_label_as_reported"].map(normalize_station)
    statistic = frame["statistic_as_reported"].str.lower().str.replace(" ", "", regex=False)
    averages = frame[statistic.eq("average")][
        ["report_month", "parameter", "station", "value"]
    ].rename(columns={"value": "station_average"})
    capture = frame[statistic.str.contains("datacapture", na=False)][
        ["report_month", "parameter", "station", "value"]
    ].rename(columns={"value": "capture_pct"})
    merged = averages.merge(capture, on=["report_month", "parameter", "station"], how="left")

    rows: list[dict[str, Any]] = []
    for (month, parameter), group in merged.groupby(["report_month", "parameter"]):
        valid_weights = group.dropna(subset=["capture_pct"])
        weighted = (
            np.average(valid_weights["station_average"], weights=valid_weights["capture_pct"])
            if not valid_weights.empty and valid_weights["capture_pct"].sum() > 0
            else np.nan
        )
        barc = group.loc[group["station"].eq("BARC/Farmgate"), "station_average"]
        rows.append(
            {
                "month_start": month,
                "variable": "pm25" if parameter == "PM2.5" else "pm10",
                "unweighted_station_mean": group["station_average"].mean(),
                "capture_weighted_mean": weighted,
                "barc_fixed_station_mean": barc.mean() if not barc.empty else np.nan,
                "station_count": group["station"].nunique(),
                "mean_capture_pct": group["capture_pct"].mean(),
            }
        )
    return pd.DataFrame(rows)


def station_tables(monthly: pd.DataFrame) -> dict[str, pd.DataFrame]:
    frame = monthly.copy()
    frame["report_month"] = pd.to_datetime(frame["report_month"])
    frame = frame[
        frame["report_month"].dt.year.between(2013, 2019)
        | frame["report_month"].dt.year.between(2022, 2025)
    ].copy()
    frame = frame[frame["parameter"].isin(["PM2.5", "PM10"])]
    frame["station"] = frame["station_label_as_reported"].map(normalize_station)
    statistic = frame["statistic_as_reported"].str.lower().str.replace(" ", "", regex=False)
    averages = frame[statistic.eq("average")][
        ["report_month", "parameter", "station", "value"]
    ].rename(columns={"value": "station_average"})
    capture = frame[statistic.str.contains("datacapture", na=False)][
        ["report_month", "parameter", "station", "value"]
    ].rename(columns={"value": "capture_pct"})
    records = averages.merge(capture, on=["report_month", "parameter", "station"], how="left")
    records["year"] = records["report_month"].dt.year
    records["month"] = records["report_month"].dt.month
    records["season"] = records["month"].map(season_for_month)
    records["era"] = np.where(records["year"].le(2019), "2013-2019", "2022-2025")

    summaries = (
        records.groupby(["era", "parameter", "station"], as_index=False)
        .agg(
            n_months=("station_average", "count"),
            first_month=("report_month", "min"),
            last_month=("report_month", "max"),
            mean=("station_average", "mean"),
            median=("station_average", "median"),
            standard_deviation=("station_average", "std"),
            mean_capture_pct=("capture_pct", "mean"),
        )
        .sort_values(["era", "parameter", "station"])
    )

    contrast_rows: list[dict[str, Any]] = []
    for (era, parameter), group in records.groupby(["era", "parameter"]):
        pivot = group.pivot_table(
            index="report_month", columns="station", values="station_average", aggfunc="mean"
        )
        stations = sorted(pivot.columns)
        for left_index, left in enumerate(stations):
            for right in stations[left_index + 1 :]:
                paired = pivot[[left, right]].dropna()
                differences = paired[left] - paired[right]
                low, high = bootstrap_mean_ci(differences, 20260807 + len(contrast_rows))
                contrast_rows.append(
                    {
                        "era": era,
                        "parameter": parameter,
                        "station_left": left,
                        "station_right": right,
                        "n_paired_months": len(paired),
                        "mean_paired_difference_left_minus_right": differences.mean(),
                        "bootstrap_ci_low": low,
                        "bootstrap_ci_high": high,
                        "spearman_rho": paired[left].corr(paired[right], method="spearman"),
                    }
                )
    return {
        "station_observations": records,
        "station_summary": summaries,
        "station_paired_contrasts": pd.DataFrame(contrast_rows),
    }


def adjusted_hac_trend(
    dates: pd.Series, values: pd.Series, start_year: int = 2022
) -> dict[str, float]:
    frame = pd.DataFrame({"date": dates, "value": values}).dropna()
    frame = frame[frame["date"].dt.year.ge(start_year)]
    frame["time_years"] = (frame["date"] - frame["date"].min()).dt.days / 365.25
    dummies = pd.get_dummies(frame["date"].dt.month.astype(str), drop_first=True, dtype=float)
    design = sm.add_constant(
        pd.concat(
            [frame[["time_years"]].reset_index(drop=True), dummies.reset_index(drop=True)], axis=1
        )
    )
    model = sm.OLS(frame["value"].to_numpy(), design).fit(cov_type="HAC", cov_kwds={"maxlags": 3})
    ci = model.conf_int().loc["time_years"]
    return {
        "n_months": len(frame),
        "slope_per_year": model.params["time_years"],
        "ci_low": ci.iloc[0],
        "ci_high": ci.iloc[1],
        "p_value": model.pvalues["time_years"],
    }


def theil_sen_trend(dates: pd.Series, values: pd.Series) -> dict[str, float]:
    frame = pd.DataFrame({"date": dates, "value": values}).dropna()
    frame = frame[frame["date"].dt.year.between(2022, 2025)].copy()
    frame["anomaly"] = frame["value"] - frame.groupby(frame["date"].dt.month)["value"].transform(
        "mean"
    )
    time_years = (frame["date"] - frame["date"].min()).dt.days / 365.25
    slope, _, low, high = stats.theilslopes(frame["anomaly"], time_years, alpha=0.95)
    return {
        "n_months": len(frame),
        "slope_per_year": slope,
        "ci_low": low,
        "ci_high": high,
        "p_value": np.nan,
    }


def trend_sensitivity_table(wide: pd.DataFrame, station_series: pd.DataFrame) -> pd.DataFrame:
    recent = wide[wide["year"].between(2022, 2025)]
    historical = wide[wide["year"].between(2013, 2019)]
    historical_without_partial_2019 = wide[wide["year"].between(2013, 2018)]
    rows: list[dict[str, Any]] = []
    for key, (column, _, _) in CORE_VARIABLES.items():
        if key in {"pm25", "pm10"}:
            for specification, subset in [
                ("historical HAC, 2013-2019", historical),
                ("historical HAC, 2013-2018", historical_without_partial_2019),
            ]:
                rows.append(
                    {
                        "variable": key,
                        "series": "city monthly mean",
                        "analysis_era": "historical",
                        "specification": specification,
                        "unit_per_year": PLAIN_UNITS[key],
                        **adjusted_hac_trend(
                            subset["month_start"], subset[column], int(subset["year"].min())
                        ),
                    }
                )
        specifications = [
            (
                "month-adjusted HAC, 2022-2025",
                adjusted_hac_trend(recent["month_start"], recent[column], 2022),
            ),
            (
                "month-adjusted HAC, 2023-2025",
                adjusted_hac_trend(recent["month_start"], recent[column], 2023),
            ),
            (
                "Theil-Sen on monthly anomalies",
                theil_sen_trend(recent["month_start"], recent[column]),
            ),
        ]
        for specification, result in specifications:
            rows.append(
                {
                    "variable": key,
                    "series": "city monthly mean",
                    "analysis_era": "recent",
                    "specification": specification,
                    "unit_per_year": PLAIN_UNITS[key],
                    **result,
                }
            )

        if key in {"pm25", "pm10"}:
            subset = station_series[
                station_series["variable"].eq(key)
                & station_series["month_start"].dt.year.between(2022, 2025)
            ]
            for series_column, series_label in [
                ("capture_weighted_mean", "capture-weighted stations"),
                ("barc_fixed_station_mean", "BARC fixed station"),
            ]:
                result = adjusted_hac_trend(subset["month_start"], subset[series_column], 2022)
                rows.append(
                    {
                        "variable": key,
                        "series": series_label,
                        "analysis_era": "recent",
                        "specification": "month-adjusted HAC, 2022-2025",
                        "unit_per_year": PLAIN_UNITS[key],
                        **result,
                    }
                )
    frame = pd.DataFrame(rows)
    frame["q_value_bh"] = fdr(frame["p_value"])
    return frame


def correlation_long(frame: pd.DataFrame, columns: dict[str, str], scope: str) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    keys = list(columns)
    for left_index, left in enumerate(keys):
        for right in keys[left_index + 1 :]:
            paired = frame[[columns[left], columns[right]]].dropna()
            rho, p_value = stats.spearmanr(paired.iloc[:, 0], paired.iloc[:, 1])
            rows.append(
                {
                    "scope": scope,
                    "variable_1": left,
                    "variable_2": right,
                    "n_paired_months": len(paired),
                    "spearman_rho": rho,
                    "p_value": p_value,
                }
            )
    result = pd.DataFrame(rows)
    result["q_value_bh"] = fdr(result["p_value"])
    return result


def multipollutant_tables(wide: pd.DataFrame) -> dict[str, pd.DataFrame]:
    historical = wide[wide["year"].between(2013, 2019)].copy()
    recent = wide[wide["year"].between(2022, 2025)].copy()
    summary_rows: list[dict[str, Any]] = []
    seasonal_rows: list[dict[str, Any]] = []
    trend_rows: list[dict[str, Any]] = []

    for key, (column, label, unit) in HISTORICAL_POLLUTANTS.items():
        resolved = historical[historical[f"{key}_unit_status"].eq("resolved")].copy()
        values = resolved[column].dropna()
        summary_rows.append(
            {
                "variable": key,
                "label": label,
                "unit": unit,
                "n_months": len(values),
                "mean": values.mean(),
                "standard_deviation": values.std(ddof=1),
                "median": values.median(),
                "q1": values.quantile(0.25),
                "q3": values.quantile(0.75),
                "minimum": values.min(),
                "maximum": values.max(),
            }
        )
        for season in SEASONS:
            season_values = resolved.loc[resolved["season"].eq(season), column].dropna()
            seasonal_rows.append(
                {
                    "variable": key,
                    "label": label,
                    "unit": unit,
                    "season": season,
                    "n_months": len(season_values),
                    "mean": season_values.mean(),
                    "median": season_values.median(),
                    "q1": season_values.quantile(0.25),
                    "q3": season_values.quantile(0.75),
                }
            )
        trend_rows.append(
            {
                "variable": key,
                "label": label,
                "unit_per_year": f"{unit}/year",
                **adjusted_hac_trend(resolved["month_start"], resolved[column], 2013),
            }
        )

    historical_columns = {key: column for key, (column, _, _) in HISTORICAL_POLLUTANTS.items()}
    recent_columns = {key: CORE_VARIABLES[key][0] for key in ("aqi", "pm25", "pm10")}
    historical_correlations = correlation_long(
        historical, historical_columns, "Historical resolved-unit months, 2013-2019"
    )
    recent_correlations = correlation_long(
        recent, recent_columns, "Recent joint monthly window, 2022-2025"
    )
    trend_frame = pd.DataFrame(trend_rows)
    trend_frame["q_value_bh"] = fdr(trend_frame["p_value"])
    return {
        "historical_pollutants": pd.DataFrame(summary_rows),
        "historical_seasonal": pd.DataFrame(seasonal_rows),
        "historical_pollutant_trends": trend_frame,
        "historical_correlations": historical_correlations,
        "recent_correlations": recent_correlations,
    }


def context_tables(
    population: pd.DataFrame,
    worldometer: pd.DataFrame,
    forest: pd.DataFrame,
    hdi: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    context = population[
        ["year", "total_population", "urban_population", "urban_share_fraction"]
    ].merge(
        hdi[["year", "hdi_undp_same_year"]],
        on="year",
        how="left",
    )
    context = context.merge(forest[["year", "tree_cover_loss_ha"]], on="year", how="left")

    summary_rows = []
    for label, column, unit in [
        ("Total population", "total_population", "persons"),
        ("Urban population", "urban_population", "persons"),
        ("Urban share", "urban_share_fraction", "fraction"),
        ("UNDP HDI", "hdi_undp_same_year", "index"),
    ]:
        values = context[["year", column]].dropna()
        start = values.iloc[0]
        end = values.iloc[-1]
        summary_rows.append(
            {
                "indicator": label,
                "unit": unit,
                "first_year": int(start["year"]),
                "first_value": start[column],
                "last_year": int(end["year"]),
                "last_value": end[column],
                "absolute_change": end[column] - start[column],
                "relative_change_pct": 100 * (end[column] / start[column] - 1),
                "scope": "Bangladesh national",
            }
        )
    forest_period = forest[forest["year"].between(2013, 2024)]
    summary_rows.append(
        {
            "indicator": "Tree-cover loss, cumulative",
            "unit": "hectares",
            "first_year": 2013,
            "first_value": forest_period.iloc[0]["tree_cover_loss_ha"],
            "last_year": 2024,
            "last_value": forest_period.iloc[-1]["tree_cover_loss_ha"],
            "absolute_change": forest_period["tree_cover_loss_ha"].sum(),
            "relative_change_pct": np.nan,
            "scope": "Bangladesh national; all-cause tree-cover loss",
        }
    )

    overlap = population.merge(worldometer, on="year", suffixes=("_un", "_worldometer"))
    overlap["total_population_difference"] = (
        overlap["total_population_worldometer"] - overlap["total_population_un"]
    )
    overlap["total_population_relative_difference_pct"] = (
        100 * overlap["total_population_difference"] / overlap["total_population_un"]
    )
    comparison = overlap[
        [
            "year",
            "total_population_un",
            "total_population_worldometer",
            "total_population_difference",
            "total_population_relative_difference_pct",
            "urban_population_un",
            "urban_population_worldometer",
        ]
    ].copy()
    comparison["interpretation"] = (
        "Total population cross-check only; urban definitions differ and are not compared analytically."
    )
    return {
        "context_trajectory": context,
        "context_change_summary": pd.DataFrame(summary_rows),
        "population_source_comparison": comparison,
    }


def data_quality_table(wide: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for key, (column, _, _) in CORE_VARIABLES.items():
        available = wide.loc[wide[column].notna(), "month_start"]
        rows.append(
            {
                "variable": key,
                "available_months": len(available),
                "first_month": available.min().date(),
                "last_month": available.max().date(),
                "complete_months_2022_2025": wide.loc[wide["year"].between(2022, 2025), column]
                .notna()
                .sum(),
                "expected_months_2022_2025": 48,
                "available_months_2013_2019": wide.loc[wide["year"].between(2013, 2019), column]
                .notna()
                .sum(),
                "expected_months_2013_2019": 84 if key in {"pm25", "pm10"} else np.nan,
            }
        )
    return pd.DataFrame(rows)


def source_qa_tables() -> tuple[pd.DataFrame, pd.DataFrame]:
    qa = pd.read_csv(PROCESSED / "doe_qa_issues.csv")
    manifest = pd.read_csv(PROCESSED / "doe_source_manifest.csv")
    qa_summary = (
        qa.groupby(["stage", "severity"], dropna=False).size().rename("issue_count").reset_index()
    )
    manifest_summary = (
        manifest.groupby(["source_kind", "extraction_status", "download_status"], dropna=False)
        .size()
        .rename("report_count")
        .reset_index()
    )
    return qa_summary, manifest_summary


def context_notes(
    population: pd.DataFrame, worldometer: pd.DataFrame, forest: pd.DataFrame
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "dataset": "UN WUP population",
                "scope": "Bangladesh national",
                "coverage": f"{population.year.min()}-{population.year.max()}",
                "research_role": "descriptive context",
                "interpretation_limit": "Not a Dhaka exposure denominator and not entered into pollution models.",
            },
            {
                "dataset": "Worldometer population",
                "scope": "Bangladesh national",
                "coverage": f"selected years {worldometer.year.min()}-{worldometer.year.max()}",
                "research_role": "cross-check of UN-derived estimates",
                "interpretation_limit": "Sparse presentation series; not an independent population source or causal predictor.",
            },
            {
                "dataset": "Global Forest Watch tree-cover loss",
                "scope": "Bangladesh national",
                "coverage": f"{forest.year.min()}-{forest.year.max()}",
                "research_role": "descriptive environmental context",
                "interpretation_limit": "Tree-cover loss is not necessarily permanent deforestation and is spatially mismatched to Dhaka AQI.",
            },
        ]
    )


def set_plot_style() -> None:
    plt.rcParams.update(
        {
            "font.size": 9,
            "axes.titlesize": 11,
            "axes.labelsize": 9,
            "figure.dpi": 120,
            "axes.spines.top": False,
            "axes.spines.right": False,
            "axes.grid": True,
            "grid.alpha": 0.2,
        }
    )


def figure_coverage(wide: pd.DataFrame) -> None:
    matrix = np.vstack(
        [wide[column].notna().to_numpy() for column, _, _ in CORE_VARIABLES.values()]
    )
    fig, axes = plt.subplots(
        3,
        1,
        figsize=(11, 9),
        gridspec_kw={"height_ratios": [0.8, 1.5, 1.35]},
        constrained_layout=True,
    )
    axes[0].imshow(matrix, aspect="auto", interpolation="nearest", cmap="Blues", vmin=0, vmax=1)
    axes[0].set_yticks(
        range(3),
        [r"$\mathrm{AQI}$", r"$\mathrm{PM}_{2.5}$", r"$\mathrm{PM}_{10}$"],
    )
    year_positions = wide.groupby("year").head(1).index.to_numpy()
    axes[0].set_xticks(year_positions, wide.loc[year_positions, "year"].astype(str), rotation=45)
    axes[0].set_title("A. Official monthly data availability")
    axes[0].set_xlabel("Blue = reported; white = unavailable")

    particulate = wide[wide["year"].between(2013, 2019) | wide["year"].between(2022, 2025)]
    for key in ("pm25", "pm10"):
        for era_index, (_, (start_year, end_year)) in enumerate(PARTICULATE_ERAS.items()):
            era = particulate[particulate["year"].between(start_year, end_year)]
            axes[1].plot(
                era["month_start"],
                era[f"{key}_mean"],
                color=COLORS[key],
                linewidth=0.8,
                alpha=0.4,
                marker="o",
                markersize=2,
                label=CORE_VARIABLES[key][1] if era_index == 0 else None,
            )
            smooth = era.set_index("month_start")[f"{key}_mean"].rolling(12, min_periods=6).mean()
            axes[1].plot(smooth.index, smooth, color=COLORS[key], linewidth=2)
    axes[1].axvspan(
        pd.Timestamp("2020-01-01"), pd.Timestamp("2021-12-31"), color="#999999", alpha=0.16
    )
    axes[1].text(
        pd.Timestamp("2021-01-01"),
        axes[1].get_ylim()[1] * 0.93,
        "Public-report gap",
        ha="center",
        va="top",
        color="#555555",
    )
    axes[1].set_ylabel(r"Monthly mean ($\mu\mathrm{g}\,\mathrm{m}^{-3}$)")
    axes[1].set_title("B. Long-term particulate context; thick lines are 12-month rolling means")
    axes[1].legend(frameon=False, ncol=2)

    for key, label, color in [
        ("pm25", r"$\mathrm{PM}_{2.5}$", COLORS["pm25"]),
        ("pm10", r"$\mathrm{PM}_{10}$", COLORS["pm10"]),
    ]:
        for _, (start_year, end_year) in PARTICULATE_ERAS.items():
            era = particulate[particulate["year"].between(start_year, end_year)]
            axes[2].plot(
                era["month_start"],
                era[f"{key}_station_count"],
                color=color,
                label=f"{label} stations" if start_year == 2013 else None,
            )
    axes[2].set_ylabel("Reporting stations")
    axes[2].set_ylim(bottom=0)
    axes[2].set_title("C. Monitoring support across the two public-report eras")
    second = axes[2].twinx()
    for era_index, (_, (start_year, end_year)) in enumerate(PARTICULATE_ERAS.items()):
        era = particulate[particulate["year"].between(start_year, end_year)]
        second.plot(
            era["month_start"],
            era["pm25_mean_data_capture_pct"],
            color="#2166AC",
            alpha=0.7,
            linestyle="--",
            label=r"$\mathrm{PM}_{2.5}$ capture" if era_index == 0 else None,
        )
    second.set_ylabel("Mean data capture (%)")
    second.set_ylim(0, 105)
    lines = [line for line in axes[2].lines + second.lines if not line.get_label().startswith("_")]
    axes[2].legend(
        lines, [line.get_label() for line in lines], loc="lower left", ncol=3, frameon=False
    )
    fig.suptitle(
        "Coverage, long-term particulate context, and monitoring support",
        fontweight="bold",
    )
    fig.savefig(FIGURES / "figure_1_coverage_and_monitoring.png", dpi=240)
    plt.close(fig)


def figure_daily_burden(daily: pd.DataFrame, monthly_exceedance: pd.DataFrame) -> None:
    fig, axes = plt.subplots(
        2,
        1,
        figsize=(11, 7),
        sharex=True,
        gridspec_kw={"height_ratios": [2.1, 1]},
        constrained_layout=True,
    )
    bands = [
        (0, 50, "#00A651"),
        (50, 100, "#FFF200"),
        (100, 150, "#F7941D"),
        (150, 200, "#ED1C24"),
        (200, 300, "#92278F"),
        (300, max(500, daily.aqi.max() + 20), "#7E0023"),
    ]
    for low, high, color in bands:
        axes[0].axhspan(low, high, color=color, alpha=0.09, linewidth=0)
    axes[0].plot(
        daily["report_date"],
        daily["aqi"],
        color="#777777",
        linewidth=0.7,
        alpha=0.65,
        label=r"Daily $\mathrm{AQI}$",
    )
    rolling = daily.set_index("report_date")["aqi"].rolling("30D", min_periods=15).mean()
    axes[0].plot(rolling.index, rolling, color="#111111", linewidth=2, label="30-day mean")
    axes[0].axhline(150, color="#B2182B", linestyle="--", linewidth=1)
    axes[0].set_ylabel(r"$\mathrm{AQI}$")
    axes[0].set_title("A. Daily observations and 30-day burden")
    axes[0].legend(frameon=False, ncol=2)

    burden = monthly_exceedance[monthly_exceedance["threshold"].eq(150)]
    yerr = np.maximum(
        np.vstack(
            [
                burden["share_pct"] - burden["wilson_ci_low_pct"],
                burden["wilson_ci_high_pct"] - burden["share_pct"],
            ]
        ),
        0,
    )
    axes[1].errorbar(
        burden["month_start"],
        burden["share_pct"],
        yerr=yerr,
        color=COLORS["aqi"],
        marker="o",
        markersize=3,
        linewidth=1.2,
        capsize=2,
    )
    axes[1].set_ylabel("Reported days\n" + r"with $\mathrm{AQI}>150$ (%)")
    axes[1].set_ylim(0, 105)
    axes[1].set_title("B. Monthly high-pollution frequency with 95% Wilson intervals")
    axes[1].xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    axes[1].xaxis.set_major_formatter(mdates.DateFormatter("%b\n%Y"))
    fig.suptitle(
        r"Unified daily $\mathrm{AQI}$ burden in Dhaka, 2022–2025", fontweight="bold"
    )
    fig.savefig(FIGURES / "figure_2_daily_aqi_burden.png", dpi=240)
    plt.close(fig)


def figure_seasonal(wide: pd.DataFrame) -> None:
    recent = wide[wide["year"].between(2022, 2025)]
    rng = np.random.default_rng(20260806)
    fig, axes = plt.subplots(1, 3, figsize=(12, 4.5), constrained_layout=True)
    for axis, (key, (column, label, unit)) in zip(axes, CORE_VARIABLES.items(), strict=True):
        values = [
            recent.loc[recent["season"].eq(season), column].dropna().to_numpy()
            for season in SEASONS
        ]
        boxes = axis.boxplot(values, patch_artist=True, widths=0.58, showfliers=False)
        for box in boxes["boxes"]:
            box.set(facecolor=COLORS[key], alpha=0.28, edgecolor=COLORS[key])
        for index, group in enumerate(values, start=1):
            axis.scatter(
                rng.normal(index, 0.045, len(group)),
                group,
                s=15,
                alpha=0.7,
                color=COLORS[key],
                edgecolor="white",
                linewidth=0.25,
            )
        axis.set_xticks(range(1, 5), ["Winter", "Pre-\nmonsoon", "Monsoon", "Post-\nmonsoon"])
        axis.set_title(label)
        axis.set_ylabel(unit)
    fig.suptitle(
        r"Seasonal distributions of monthly $\mathrm{AQI}$ and particulate matter, 2022–2025",
        fontweight="bold",
    )
    fig.savefig(FIGURES / "figure_3_seasonal_particulate_burden.png", dpi=240)
    plt.close(fig)


def figure_episodes(episodes: pd.DataFrame, daily_season: pd.DataFrame) -> None:
    top = episodes.nlargest(12, ["duration_days", "peak_aqi"]).sort_values("duration_days")
    fig, axes = plt.subplots(
        1, 2, figsize=(12, 5.8), gridspec_kw={"width_ratios": [1.35, 1]}, constrained_layout=True
    )
    labels = [f"{row.start_date:%d %b %Y}" for row in top.itertuples()]
    scatter = axes[0].scatter(
        top["duration_days"], np.arange(len(top)), c=top["peak_aqi"], cmap="magma", s=60, zorder=3
    )
    axes[0].hlines(np.arange(len(top)), 0, top["duration_days"], color="#999999", linewidth=1)
    axes[0].set_yticks(np.arange(len(top)), labels)
    axes[0].set_xlabel(r"Consecutive reported days with $\mathrm{AQI}>150$")
    axes[0].set_title("A. Twelve longest high-pollution episodes")
    fig.colorbar(scatter, ax=axes[0], label=r"Episode peak $\mathrm{AQI}$", shrink=0.75)

    x = np.arange(len(SEASONS))
    width = 0.24
    for offset, threshold, color in [
        (-width, 100, "#F7941D"),
        (0, 150, "#ED1C24"),
        (width, 200, "#92278F"),
    ]:
        axes[1].bar(
            x + offset,
            daily_season[f"share_gt_{threshold}_pct"],
            width,
            label=rf"$>{threshold}$",
            color=color,
            alpha=0.85,
        )
    axes[1].set_xticks(x, ["Winter", "Pre-\nmonsoon", "Monsoon", "Post-\nmonsoon"])
    axes[1].set_ylabel("Reported days above threshold (%)")
    axes[1].set_ylim(0, 105)
    axes[1].set_title("B. Seasonal threshold burden")
    axes[1].legend(title=r"$\mathrm{AQI}$", frameon=False, ncol=3, loc="upper center")
    fig.suptitle(
        r"Persistence and seasonal concentration of high-$\mathrm{AQI}$ episodes, 2022–2025",
        fontweight="bold",
    )
    fig.savefig(FIGURES / "figure_4_pollution_episodes.png", dpi=240)
    plt.close(fig)


def figure_trends(trends: pd.DataFrame) -> None:
    labels = {
        "aqi": r"$\mathrm{AQI}$",
        "pm25": r"$\mathrm{PM}_{2.5}$",
        "pm10": r"$\mathrm{PM}_{10}$",
    }
    fig, axes = plt.subplots(1, 3, figsize=(13.5, 7.2), constrained_layout=True)
    for axis, variable in zip(axes, CORE_VARIABLES, strict=True):
        subset = trends[trends["variable"].eq(variable)].copy().reset_index(drop=True)
        subset["plot_label"] = subset["specification"].str.replace(
            "month-adjusted HAC, ", "HAC ", regex=False
        )
        subset["plot_label"] = subset["plot_label"].str.replace(
            "historical HAC, ", "Historical HAC ", regex=False
        )
        subset["plot_label"] = subset["plot_label"].replace(
            {"Theil-Sen on monthly anomalies": "Theil-Sen anomalies"}
        )
        subset.loc[subset["series"].ne("city monthly mean"), "plot_label"] = subset.loc[
            subset["series"].ne("city monthly mean"), "series"
        ]
        subset["plot_label"] = subset["plot_label"].replace(
            {
                "capture-weighted stations": "Capture-weighted",
                "BARC fixed station": "BARC fixed station",
            }
        )
        y = np.arange(len(subset))[::-1]
        axis.axvline(0, color="#555555", linewidth=1)
        axis.errorbar(
            subset["slope_per_year"],
            y,
            xerr=np.vstack(
                [
                    subset["slope_per_year"] - subset["ci_low"],
                    subset["ci_high"] - subset["slope_per_year"],
                ]
            ),
            fmt="o",
            color=COLORS[variable],
            capsize=3,
        )
        axis.set_yticks(y, subset["plot_label"])
        axis.set_title(labels[variable])
        axis.set_xlabel(r"Estimated trend, $\hat{\beta}$ per year (95% CI)")
    fig.suptitle(
        "Trend estimates across temporal and monitoring-sensitivity specifications",
        fontweight="bold",
    )
    fig.savefig(FIGURES / "figure_5_trend_sensitivity.png", dpi=240)
    plt.close(fig)


def figure_station_heterogeneity(station_observations: pd.DataFrame) -> None:
    fig, axes = plt.subplots(2, 2, figsize=(13, 8), constrained_layout=True)
    era_order = ["2013-2019", "2022-2025"]
    parameter_order = ["PM2.5", "PM10"]
    colors = {"PM2.5": COLORS["pm25"], "PM10": COLORS["pm10"]}
    for row, parameter in enumerate(parameter_order):
        for column, era in enumerate(era_order):
            axis = axes[row, column]
            subset = station_observations[
                station_observations["parameter"].eq(parameter)
                & station_observations["era"].eq(era)
            ]
            stations = sorted(subset["station"].unique())
            values = [
                subset.loc[subset["station"].eq(station), "station_average"].dropna()
                for station in stations
            ]
            boxes = axis.boxplot(values, patch_artist=True, showfliers=False)
            for box in boxes["boxes"]:
                box.set(facecolor=colors[parameter], alpha=0.28, edgecolor=colors[parameter])
            axis.set_xticks(range(1, len(stations) + 1), stations, rotation=18, ha="right")
            axis.set_ylabel(r"Monthly station mean ($\mu\mathrm{g}\,\mathrm{m}^{-3}$)")
            axis.set_title(f"{parameter}, {era}")
    fig.suptitle(
        "Station-level particulate heterogeneity within each public-report era",
        fontweight="bold",
    )
    fig.savefig(FIGURES / "figure_6_station_heterogeneity.png", dpi=240)
    plt.close(fig)


def draw_correlation_heatmap(
    axis: plt.Axes, frame: pd.DataFrame, columns: dict[str, str], title: str
) -> None:
    renamed = frame[list(columns.values())].rename(
        columns={value: key for key, value in columns.items()}
    )
    matrix = renamed.corr(method="spearman")
    image = axis.imshow(matrix, cmap="RdBu_r", vmin=-1, vmax=1)
    labels = list(matrix.columns)
    axis.set_xticks(range(len(labels)), labels, rotation=45, ha="right")
    axis.set_yticks(range(len(labels)), labels)
    for row in range(len(labels)):
        for column in range(len(labels)):
            value = matrix.iloc[row, column]
            axis.text(
                column,
                row,
                f"{value:.2f}",
                ha="center",
                va="center",
                color="white" if abs(value) > 0.55 else "#222222",
                fontsize=8,
            )
    axis.set_title(title)
    return image


def figure_multipollutant(wide: pd.DataFrame) -> None:
    historical = wide[wide["year"].between(2013, 2019)].copy()
    recent = wide[wide["year"].between(2022, 2025)].copy()
    historical_columns = {
        "PM2.5": "pm25_mean",
        "PM10": "pm10_mean",
        "NO2": "no2_mean",
        "SO2": "so2_mean",
        "CO": "co_mean",
        "O3": "o3_mean",
    }
    recent_columns = {"AQI": "aqi_mean", "PM2.5": "pm25_mean", "PM10": "pm10_mean"}
    fig = plt.figure(figsize=(14, 8.5), constrained_layout=True)
    grid = fig.add_gridspec(2, 2, height_ratios=[1, 1.05], width_ratios=[1.35, 1])
    historical_axis = fig.add_subplot(grid[0, 0])
    recent_axis = fig.add_subplot(grid[0, 1])
    seasonal_axis = fig.add_subplot(grid[1, :])
    image = draw_correlation_heatmap(
        historical_axis,
        historical,
        historical_columns,
        "A. Historical monthly Spearman correlations, 2013–2019",
    )
    draw_correlation_heatmap(
        recent_axis,
        recent,
        recent_columns,
        "B. Recent joint-window correlations, 2022–2025",
    )
    fig.colorbar(image, ax=[historical_axis, recent_axis], label=r"Spearman $\rho$", shrink=0.8)

    standardized = historical[list(historical_columns.values())].apply(
        lambda series: (series - series.mean()) / series.std(ddof=1)
    )
    standardized["season"] = historical["season"].to_numpy()
    seasonal = standardized.groupby("season").mean().reindex(SEASONS)
    x = np.arange(len(SEASONS))
    palette = plt.get_cmap("tab10")
    for index, (label, column) in enumerate(historical_columns.items()):
        seasonal_axis.plot(
            x,
            seasonal[column],
            marker="o",
            linewidth=1.8,
            color=palette(index),
            label=label,
        )
    seasonal_axis.axhline(0, color="#555555", linewidth=0.8)
    seasonal_axis.set_xticks(x, SEASONS)
    seasonal_axis.set_ylabel("Mean standardized concentration")
    seasonal_axis.set_title("C. Historical multi-pollutant seasonal profiles")
    seasonal_axis.legend(ncol=6, frameon=False, loc="upper center")
    fig.suptitle(
        "Multi-pollutant dependence and seasonal structure",
        fontweight="bold",
    )
    fig.savefig(FIGURES / "figure_7_multipollutant_structure.png", dpi=240)
    plt.close(fig)


def figure_context_framework(context: pd.DataFrame) -> None:
    fig = plt.figure(figsize=(13, 8.5), constrained_layout=True)
    grid = fig.add_gridspec(2, 2, height_ratios=[1, 1.25])
    framework = fig.add_subplot(grid[0, :])
    demographics = fig.add_subplot(grid[1, 0])
    forest_axis = fig.add_subplot(grid[1, 1])

    framework.set_xlim(0, 1)
    framework.set_ylim(0, 1)
    framework.axis("off")
    boxes = [
        (0.11, 0.62, "Population, urbanization,\nand development"),
        (0.37, 0.62, "Activity and emissions\n(traffic, industry, dust)"),
        (0.63, 0.62, "Ambient pollutant\nconcentrations"),
        (0.88, 0.62, "Published AQI\nand exposure burden"),
        (0.37, 0.16, "Forest change, fires,\nand regional transport"),
        (0.63, 0.16, "Meteorology and\nboundary-layer mixing"),
    ]
    for x, y, label in boxes:
        framework.text(
            x,
            y,
            label,
            transform=framework.transAxes,
            ha="center",
            va="center",
            bbox={"boxstyle": "round,pad=0.5", "facecolor": "#EAF2F8", "edgecolor": "#1F4E78"},
        )
    arrows = [
        ((0.22, 0.62), (0.27, 0.62)),
        ((0.48, 0.62), (0.54, 0.62)),
        ((0.74, 0.62), (0.79, 0.62)),
        ((0.46, 0.28), (0.58, 0.48)),
        ((0.63, 0.29), (0.63, 0.47)),
    ]
    for start, end in arrows:
        framework.annotate(
            "",
            xy=end,
            xytext=start,
            xycoords="axes fraction",
            arrowprops={"arrowstyle": "->", "color": "#444444", "lw": 1.4},
        )
    framework.set_title(
        "A. Conceptual framework: context informs hypotheses but does not identify causation"
    )

    indexed = context.set_index("year")
    for column, label, color in [
        ("total_population", "Total population", "#2166AC"),
        ("urban_population", "Urban population", "#1B7837"),
        ("hdi_undp_same_year", "UNDP HDI", "#762A83"),
    ]:
        values = indexed[column].dropna()
        demographics.plot(
            values.index,
            100 * values / values.iloc[0],
            label=label,
            color=color,
            marker="o",
            markersize=3,
        )
    demographics.axhline(100, color="#777777", linewidth=0.8)
    demographics.set_ylabel("Index (first available year = 100)")
    demographics.set_title("B. National demographic and development context")
    demographics.legend(frameon=False)

    forest_values = context.dropna(subset=["tree_cover_loss_ha"])
    forest_axis.bar(
        forest_values["year"],
        forest_values["tree_cover_loss_ha"] / 1_000,
        color="#8C510A",
        alpha=0.8,
    )
    forest_axis.set_ylabel("Tree-cover loss (thousand ha)")
    forest_axis.set_title("C. National all-cause tree-cover loss")
    fig.suptitle(
        "Structural context and the limits of ecological linkage",
        fontweight="bold",
    )
    fig.savefig(FIGURES / "figure_8_contextual_framework.png", dpi=240, bbox_inches="tight")
    plt.close(fig)


def write_frame(workbook: Workbook, name: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(frame.columns.tolist())
    for row in frame.itertuples(index=False, name=None):
        sheet.append([None if pd.isna(value) else value for value in row])
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(frame.columns, start=1):
        width = min(max(len(str(column)) + 2, 12), 45)
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_analysis_workbook(tables: dict[str, pd.DataFrame]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    readme = pd.DataFrame(
        [
            ("Purpose", "Paper-ready analysis of official DoE Dhaka observations."),
            (
                "Historical particulate period",
                "2013-2019, analyzed as a separate public-report era.",
            ),
            (
                "Recent joint monthly period",
                "2022-2025; 2026 is partial and excluded from estimates.",
            ),
            (
                "Archive discontinuity",
                "No linked monthly reports for 2020-2021; trends are not fitted across the gap.",
            ),
            (
                "Primary daily period",
                "2022-2025 unified from monthly-report Table 6 and the standalone archive using the month-level source selected for the monthly AQI aggregate.",
            ),
            (
                "Core outcomes",
                "AQI, PM2.5, and PM10 are primary outcomes; historical gases are analyzed only for 2013-2019 resolved-unit months.",
            ),
            (
                "Episodes",
                "Consecutive reported calendar days with AQI > 150; missing dates break episodes.",
            ),
            (
                "Trends",
                "Era-specific month-adjusted OLS with HAC errors, historical end-year and recent start-year sensitivities, Theil-Sen anomalies, capture weighting, and fixed-station checks.",
            ),
            (
                "Context",
                "Population and tree-cover loss are national descriptive context only; no ecological causal correlations are fitted.",
            ),
            (
                "Figures",
                "Eight integrated figures covering coverage, daily burden, seasonality, episodes, trends, stations, multi-pollutant structure, and contextual pathways.",
            ),
        ],
        columns=["item", "detail"],
    )
    write_frame(workbook, "read_me", readme)
    for name, frame in tables.items():
        write_frame(workbook, name[:31], frame)
    workbook.save(OUTPUT / "dhaka_doe_analysis.xlsx")


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    FIGURES.mkdir(parents=True, exist_ok=True)
    set_plot_style()

    wide = pd.read_csv(PROCESSED / "doe_monthly_dataset_wide.csv", parse_dates=["month_start"])
    daily_raw = pd.read_csv(PROCESSED / "doe_daily_dhaka_aqi.csv", parse_dates=["report_date"])
    monthly_report_aqi = pd.read_csv(
        PROCESSED / "doe_monthly_report_dhaka_aqi.csv",
        parse_dates=["report_month", "aqi_date"],
    )
    monthly = pd.read_csv(PROCESSED / "doe_monthly_dhaka.csv")
    population = pd.read_csv(CONTEXT / "bangladesh_population.csv")
    worldometer = pd.read_csv(CONTEXT / "bangladesh_population_worldometer.csv")
    forest = pd.read_csv(CONTEXT / "bangladesh_tree_cover_loss.csv")
    hdi = pd.read_csv(CONTEXT / "bangladesh_hdi.csv")

    descriptive = descriptive_table(wide)
    seasonal, season_tests = seasonal_tables(wide)
    long_term_particulate = long_term_particulate_tables(wide)
    multipollutant = multipollutant_tables(wide)
    daily = prepare_daily(daily_raw, monthly_report_aqi, wide)
    concordance = daily_source_concordance(daily_raw, monthly_report_aqi)
    daily_metadata = standalone_daily_metadata(daily_raw)
    daily_tables = daily_burden_tables(daily)
    episodes = episode_table(daily)
    station_series = station_sensitivity_series(monthly)
    stations = station_tables(monthly)
    trends = trend_sensitivity_table(wide, station_series)
    context = context_tables(population, worldometer, forest, hdi)
    qa_summary, manifest_summary = source_qa_tables()

    tables = {
        "descriptive_main": descriptive,
        "seasonal_main": seasonal,
        "season_tests": season_tests,
        **long_term_particulate,
        **multipollutant,
        **daily_tables,
        **concordance,
        **daily_metadata,
        "daily_unified": daily,
        "episodes_gt_150": episodes,
        "trend_sensitivity": trends,
        "station_sensitivity": station_series,
        **stations,
        "data_quality": data_quality_table(wide),
        "qa_summary": qa_summary,
        "manifest_summary": manifest_summary,
        "context_notes": context_notes(population, worldometer, forest),
        **context,
        "population_un": population,
        "population_worldometer": worldometer,
        "tree_cover_loss": forest,
        "hdi": hdi,
    }

    figure_coverage(wide)
    figure_daily_burden(daily, daily_tables["monthly_exceedance"])
    figure_seasonal(wide)
    figure_episodes(episodes, daily_tables["daily_season"])
    figure_trends(trends)
    figure_station_heterogeneity(stations["station_observations"])
    figure_multipollutant(wide)
    figure_context_framework(context["context_trajectory"])
    write_analysis_workbook(tables)
    print(f"Wrote {OUTPUT / 'dhaka_doe_analysis.xlsx'}")
    print("Wrote eight integrated paper figures")


if __name__ == "__main__":
    main()
