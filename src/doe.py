"""Official Bangladesh DoE archive discovery, extraction, and workbook helpers."""

from __future__ import annotations

import calendar
import hashlib
import re
import ssl
import subprocess
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DAILY_ARCHIVE_URL = "https://doe.gov.bd/pages/static-pages/6922dfba933eb65569e23b0a"
MONTHLY_ARCHIVE_URL = "https://doe.gov.bd/pages/static-pages/6922de32933eb65569e18f46"
ALLOWED_DOWNLOAD_HOST = "objectstorage.ap-dcc-gazipur-1.oraclecloud15.com"
USER_AGENT = "dhaka-doe-aqi-research/1.0 (reproducible public-data extraction)"

MONTHLY_YEAR_PAGES = {
    2026: "https://doe.gov.bd/pages/static-pages/69ce0b957708a43a552df023",
    2025: "https://doe.gov.bd/pages/static-pages/6922db96933eb65569e0b11a",
    2024: "https://doe.gov.bd/pages/static-pages/6922dcba933eb65569e11af7",
    2023: "https://doe.gov.bd/pages/static-pages/6922db6c933eb65569e0a11e",
    2022: "https://doe.gov.bd/pages/static-pages/6922df86933eb65569e228da",
    2019: "https://doe.gov.bd/pages/static-pages/6922df19933eb65569e1fd66",
    2018: "https://doe.gov.bd/pages/static-pages/6922dce3933eb65569e1296a",
    2017: "https://doe.gov.bd/pages/static-pages/6922dd4a933eb65569e14568",
    2016: "https://doe.gov.bd/pages/static-pages/6922dd14933eb65569e136b6",
    2015: "https://doe.gov.bd/pages/static-pages/6922dc8d933eb65569e10f17",
    2014: "https://doe.gov.bd/pages/static-pages/6922e102933eb65569e298e8",
    2013: "https://doe.gov.bd/pages/static-pages/6922e091933eb65569e27a10",
}

MONTHS = {
    name.lower(): number
    for number, name in enumerate(
        [
            "January",
            "February",
            "March",
            "April",
            "May",
            "June",
            "July",
            "August",
            "September",
            "October",
            "November",
            "December",
        ],
        start=1,
    )
}


@dataclass(frozen=True)
class ArchiveRecord:
    source_kind: str
    period: str
    label: str
    url: str
    archive_page: str


class TableHTMLParser(HTMLParser):
    """Collect rendered HTML table rows and their links using the standard library."""

    def __init__(self) -> None:
        super().__init__()
        self.in_row = False
        self.in_cell = False
        self.current_cell = ""
        self.cells: list[str] = []
        self.hrefs: list[str] = []
        self.rows: list[tuple[list[str], list[str]]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        if tag == "tr":
            self.in_row = True
            self.cells = []
            self.hrefs = []
        elif tag in {"td", "th"} and self.in_row:
            self.in_cell = True
            self.current_cell = ""
        elif tag == "a" and self.in_row and attributes.get("href"):
            self.hrefs.append(str(attributes["href"]))

    def handle_data(self, data: str) -> None:
        if self.in_cell:
            self.current_cell += data

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self.in_cell:
            self.cells.append(clean_text(self.current_cell))
            self.in_cell = False
        elif tag == "tr" and self.in_row:
            self.rows.append((self.cells, self.hrefs))
            self.in_row = False


class LinkHTMLParser(HTMLParser):
    """Collect links and their visible text from an archive landing page."""

    def __init__(self) -> None:
        super().__init__()
        self.current_href: str | None = None
        self.current_text: list[str] = []
        self.links: list[tuple[str, str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag != "a":
            return
        href = dict(attrs).get("href")
        if href:
            self.current_href = href
            self.current_text = []

    def handle_data(self, data: str) -> None:
        if self.current_href is not None:
            self.current_text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "a" and self.current_href is not None:
            self.links.append((clean_text(" ".join(self.current_text)), self.current_href))
            self.current_href = None
            self.current_text = []


def _normalize_archive_url(url: str) -> str:
    """Repair the archive's occasional exact double-paste of an attachment URL."""
    second = url.find("https://", len("https://"))
    if second > 0 and url[:second] == url[second:]:
        return url[:second]
    return url


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def make_ssl_context(intermediate_certificate: Path) -> ssl.SSLContext:
    """Use normal system roots plus DoE's currently omitted issuing certificate."""
    context = ssl.create_default_context()
    context.load_verify_locations(cafile=intermediate_certificate)
    context.check_hostname = True
    context.verify_mode = ssl.CERT_REQUIRED
    return context


def fetch_bytes(url: str, context: ssl.SSLContext, retries: int = 4) -> bytes:
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme != "https" or parsed.hostname not in {"doe.gov.bd", ALLOWED_DOWNLOAD_HOST}:
        raise ValueError(f"Refusing unapproved URL: {url}")
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    last_error: Exception | None = None
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(request, context=context, timeout=120) as response:
                return response.read()
        except Exception as error:  # network errors vary across Python/OpenSSL versions
            last_error = error
            if attempt + 1 < retries:
                time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"Failed after {retries} attempts: {url}: {last_error}")


def discover_daily(html: str) -> list[ArchiveRecord]:
    parser = TableHTMLParser()
    parser.feed(html)
    records: list[ArchiveRecord] = []
    for cells, hrefs in parser.rows:
        dates = [cell for cell in cells if re.fullmatch(r"\d{2}/\d{2}/\d{4}", cell)]
        urls = [
            normalized
            for url in hrefs
            if urllib.parse.urlparse(normalized := _normalize_archive_url(url)).hostname
            == ALLOWED_DOWNLOAD_HOST
            and re.search(r"\.(pdf|docx)$", normalized, flags=re.IGNORECASE)
        ]
        if not dates or not urls:
            continue
        date = datetime.strptime(dates[0], "%d/%m/%Y").date().isoformat()
        records.append(
            ArchiveRecord("daily_aqi", date, cells[0] if cells else date, urls[0], DAILY_ARCHIVE_URL)
        )
    unique: dict[tuple[str, str], ArchiveRecord] = {}
    for record in records:
        unique[(record.period, record.url)] = record
    return sorted(unique.values(), key=lambda record: (record.period, record.url))


def discover_monthly_year_pages(html: str) -> dict[int, str]:
    """Discover year-specific report pages linked by the monthly master archive."""
    parser = LinkHTMLParser()
    parser.feed(html)
    pages: dict[int, str] = {}
    for label, href in parser.links:
        match = re.search(r"Monthly\s+Air\s+Quality\s+Report\s+(20\d{2})", label, re.IGNORECASE)
        if match is None:
            continue
        url = urllib.parse.urljoin(MONTHLY_ARCHIVE_URL, href)
        parsed = urllib.parse.urlparse(url)
        if (
            parsed.scheme == "https"
            and parsed.hostname == "doe.gov.bd"
            and parsed.path.startswith("/pages/static-pages/")
        ):
            pages[int(match.group(1))] = url
    return dict(sorted(pages.items(), reverse=True))


def discover_monthly(year: int, archive_page: str, html: str) -> list[ArchiveRecord]:
    parser = TableHTMLParser()
    parser.feed(html)
    records: list[ArchiveRecord] = []
    for cells, hrefs in parser.rows:
        label = " ".join(cells)
        match = re.search(
            r"(January|February|March|April|May|June|July|August|September|October|November|December)(?:\s+(20\d{2}))?",
            label,
            flags=re.IGNORECASE,
        )
        urls = [
            normalized
            for url in hrefs
            if urllib.parse.urlparse(normalized := _normalize_archive_url(url)).hostname
            == ALLOWED_DOWNLOAD_HOST
            and re.search(r"\.(pdf|docx)$", normalized, flags=re.IGNORECASE)
        ]
        if not match or not urls:
            continue
        reported_year = int(match.group(2)) if match.group(2) else year
        if reported_year != year:
            continue
        month = MONTHS[match.group(1).lower()]
        period = f"{year:04d}-{month:02d}"
        reported_label = clean_text(match.group(0))
        if match.group(2) is None:
            reported_label = f"{reported_label} {year} (year supplied by archive page)"
        records.append(ArchiveRecord("monthly_report", period, reported_label, urls[0], archive_page))
    unique = {(record.period, record.url): record for record in records}
    return sorted(unique.values(), key=lambda record: (record.period, record.url))


def attachment_path(root: Path, record: ArchiveRecord) -> Path:
    suffix = Path(urllib.parse.urlparse(record.url).path).suffix.lower()
    basename = Path(urllib.parse.urlparse(record.url).path).stem
    safe_period = record.period.replace("-", "_")
    return root / record.source_kind / f"{safe_period}_{basename}{suffix}"


def save_attachment(root: Path, record: ArchiveRecord, context: ssl.SSLContext) -> dict[str, Any]:
    path = attachment_path(root, record)
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        content = fetch_bytes(record.url, context)
        temporary = path.with_suffix(path.suffix + ".part")
        temporary.write_bytes(content)
        temporary.replace(path)
    content = path.read_bytes()
    suffix = path.suffix.lower()
    valid_signature = (suffix == ".pdf" and content.startswith(b"%PDF")) or (
        suffix == ".docx" and content.startswith(b"PK")
    )
    if not valid_signature:
        raise ValueError(f"Unexpected file signature for {path}")
    return {
        "source_kind": record.source_kind,
        "period": record.period,
        "source_label": record.label,
        "archive_page": record.archive_page,
        "source_url": record.url,
        "local_path": path.as_posix(),
        "file_type": suffix.removeprefix("."),
        "file_size": len(content),
        "sha256": hashlib.sha256(content).hexdigest(),
        "download_status": "ok",
    }


def _normalized_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean_text(value).lower()).strip()


def _find_daily_row(tables: list[list[list[Any]]]) -> tuple[list[Any], list[Any]] | None:
    for table in tables:
        for header_index, header in enumerate(table):
            normalized = [_normalized_header(cell) for cell in header]
            if not any("city" in cell for cell in normalized):
                continue
            if not any(cell == "aqi" or "aqi value" in cell for cell in normalized):
                continue
            for row in table[header_index + 1 :]:
                if row and re.match(r"^dhaka", clean_text(row[0]), flags=re.IGNORECASE):
                    return header, row
    return None


def _column_index(header: list[Any], predicate: Any) -> int | None:
    for index, value in enumerate(header):
        if predicate(_normalized_header(value)):
            return index
    return None


def _daily_values(header: list[Any], row: list[Any]) -> dict[str, Any]:
    city_index = _column_index(header, lambda value: "city" in value)
    aqi_index = _column_index(header, lambda value: value == "aqi" or "aqi value" in value)
    pollutant_index = _column_index(header, lambda value: "pollutant" in value)
    category_index = _column_index(header, lambda value: "category" in value)
    comments_index = _column_index(header, lambda value: "comment" in value)
    if city_index is None or aqi_index is None:
        raise ValueError("Daily table lacks City/AQI columns")

    def get(index: int | None) -> str:
        return clean_text(row[index]) if index is not None and index < len(row) else ""

    raw_aqi = get(aqi_index)
    numeric_match = re.fullmatch(r"\d+(?:\.0+)?", raw_aqi.replace(",", ""))
    return {
        "city_as_reported": get(city_index),
        "aqi_as_reported": raw_aqi,
        "aqi": int(float(raw_aqi.replace(",", ""))) if numeric_match else None,
        "responsible_pollutant": get(pollutant_index),
        "aqi_category_as_reported": get(category_index),
        "comments": get(comments_index),
    }


def _scheme_from_text(text: str) -> str:
    lowered = text.lower()
    if "caution" in lowered or "extremely unhealthy" in lowered:
        return "bangladesh_legacy_caution_extremely_unhealthy"
    if "unhealthy for sensitive" in lowered or "hazardous" in lowered:
        return "bangladesh_current_sensitive_group_hazardous"
    return "scheme_not_resolved"


def _metadata_from_text(text: str) -> dict[str, str]:
    published = re.search(
        r"Published(?: Date)? (?:on|in) Website\s*(?::|on)?\s*(\d{1,2}/\d{1,2}/\d{4})",
        text,
        flags=re.IGNORECASE,
    )
    aqi_date = re.search(r"AQI Date\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", text, flags=re.IGNORECASE)
    notes_match = re.search(r"Note\s*:(.*?)(?:AQI Scheme|$)", text, flags=re.IGNORECASE | re.DOTALL)

    def iso(match: re.Match[str] | None) -> str:
        if match is None:
            return ""
        return datetime.strptime(match.group(1), "%d/%m/%Y").date().isoformat()

    return {
        "published_date": iso(published),
        "aqi_date_in_document": iso(aqi_date),
        "dhaka_basis_note": clean_text(notes_match.group(1)) if notes_match else "",
        "source_category_scheme": _scheme_from_text(text),
    }


def _docx_tables_and_text(path: Path) -> tuple[list[list[list[str]]], str]:
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with ZipFile(path) as archive:
        root = ET.fromstring(archive.read("word/document.xml"))
    tables: list[list[list[str]]] = []
    for table in root.findall(".//w:tbl", namespace):
        rows: list[list[str]] = []
        for row in table.findall("./w:tr", namespace):
            cells: list[str] = []
            for cell in row.findall("./w:tc", namespace):
                text = " ".join(
                    clean_text(node.text)
                    for node in cell.findall(".//w:t", namespace)
                    if clean_text(node.text)
                )
                cells.append(clean_text(text))
            rows.append(cells)
        tables.append(rows)
    paragraphs = []
    for paragraph in root.findall(".//w:p", namespace):
        value = "".join(node.text or "" for node in paragraph.findall(".//w:t", namespace))
        if clean_text(value):
            paragraphs.append(clean_text(value))
    return tables, "\n".join(paragraphs)


def extract_daily(path: Path, source: dict[str, Any]) -> dict[str, Any]:
    if path.suffix.lower() == ".pdf":
        tables: list[list[list[Any]]] = []
        text_parts: list[str] = []
        with pdfplumber.open(path) as document:
            for page in document.pages:
                text_parts.append(page.extract_text() or "")
                tables.extend(page.extract_tables())
        method = "pdfplumber_table"
        text = "\n".join(text_parts)
    elif path.suffix.lower() == ".docx":
        tables, text = _docx_tables_and_text(path)
        method = "docx_xml_table"
    else:
        raise ValueError(f"Unsupported daily attachment: {path}")
    found = _find_daily_row(tables)
    if found is None:
        raise ValueError("No Dhaka row found in daily AQI table")
    values = _daily_values(*found)
    metadata = _metadata_from_text(text)
    document_date = metadata["aqi_date_in_document"] or source["period"]
    return {
        "report_date": source["period"],
        "document_aqi_date": document_date,
        "published_date": metadata["published_date"],
        "city": "Dhaka",
        **values,
        "is_missing": values["aqi"] is None,
        "source_category_scheme": metadata["source_category_scheme"],
        "dhaka_basis_note": metadata["dhaka_basis_note"],
        "extraction_method": method,
        "source_url": source["source_url"],
        "source_sha256": source["sha256"],
        "source_file_type": source["file_type"],
        "selected_record": False,
        "duplicate_date": False,
        "qa_status": "ok" if document_date == source["period"] else "document_date_mismatch",
    }


def _is_dhaka_station(value: str) -> bool:
    normalized = _normalized_header(value)
    patterns = [
        r"^doe$",
        r"barc",
        r"farmgate",
        r"darus|d salam",
        r"s bhaban|sangshad",
        r"buet",
        r"nagor|nagar",
        r"^dhaka$",
        r"cams 1\b|cams 2\b|cams 3\b|c cams 25\b|c cams 31\b",
    ]
    return any(re.search(pattern, normalized) for pattern in patterns)


def _reported_value(value: str) -> float | None:
    cleaned = clean_text(value).replace(",", "")
    if cleaned.upper() in {"", "DNA", "NA", "N/A", "NIL", "-", "--"}:
        return None
    return float(cleaned) if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", cleaned) else None


def _reported_unit(statistic: str) -> str:
    normalized = statistic.lower()
    if "capture" in normalized or "%" in statistic:
        return "percent"
    if "day" in normalized:
        return "days"
    if "hour" in normalized:
        return "hours"
    return "not_stated_in_summary_table"


def _compact(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _pollutant_identity(value: str) -> str | None:
    compact = _compact(value)
    if "pm25" in compact or re.match(r"^pm.*24hr25", compact):
        return "PM2.5"
    if "pm10" in compact or re.match(r"^pm.*24hr10", compact):
        return "PM10"
    if "so2" in compact or re.match(r"^so.*24hr2", compact):
        return "SO2"
    if "no2" in compact or re.match(r"^no.*24hr2", compact):
        return "NO2"
    if re.search(r"(^|[^a-z])co([^a-z]|$)", value.lower()) or compact.startswith("co8"):
        return "CO"
    if "o3" in compact or re.match(r"^o.*8hr3", compact):
        return "O3"
    return None


def _pollutant_unit(value: str) -> str:
    compact = _compact(value)
    if "ugm3" in compact or "µgm3" in compact:
        return "ug/m3"
    if "ppb" in compact:
        return "ppb"
    if "ppm" in compact:
        return "ppm"
    return "not_resolved"


def _char_lines(page: Any, tolerance: float = 1.0) -> list[tuple[float, list[dict[str, Any]]]]:
    """Group PDF characters into visual lines while retaining horizontal positions."""
    lines: list[tuple[float, list[dict[str, Any]]]] = []
    for character in sorted(page.chars, key=lambda item: (float(item["top"]), float(item["x0"]))):
        top = float(character["top"])
        if not lines or abs(top - lines[-1][0]) > tolerance:
            lines.append((top, [character]))
        else:
            lines[-1][1].append(character)
    for _, characters in lines:
        characters.sort(key=lambda item: float(item["x0"]))
    return lines


def _chars_text(characters: list[dict[str, Any]], left: float, right: float) -> str:
    return clean_text(
        "".join(
            str(character.get("text", ""))
            for character in characters
            if left <= (float(character["x0"]) + float(character["x1"])) / 2 < right
        )
    )


def _header_span_center(characters: list[dict[str, Any]], needle: str) -> float | None:
    searchable: list[tuple[str, dict[str, Any]]] = []
    for character in characters:
        for token in re.sub(r"[^a-z0-9]", "", str(character.get("text", "")).lower()):
            searchable.append((token, character))
    joined = "".join(token for token, _ in searchable)
    start = joined.find(needle)
    if start < 0:
        return None
    matched = [character for _, character in searchable[start : start + len(needle)]]
    return (min(float(item["x0"]) for item in matched) + max(float(item["x1"]) for item in matched)) / 2


def _extract_monthly_positional(page: Any, page_number: int, source: dict[str, Any]) -> list[dict[str, Any]]:
    """Fallback for DoE summary pages whose borderless tables defeat line detection."""
    text = page.extract_text() or ""
    if "summary of components" not in text.lower():
        return []
    lines = _char_lines(page)
    header: tuple[float, list[dict[str, Any]]] | None = None
    for candidate in lines:
        compact = _compact(_chars_text(candidate[1], 0, page.width))
        if "parameter" in compact and "summary" in compact:
            header = candidate
            break
    if header is None:
        return []

    header_band = [
        candidate for candidate in lines if header[0] - 30 <= candidate[0] <= header[0] + 35
    ]

    def band_center(needle: str) -> float | None:
        for _, characters in header_band:
            center = _header_span_center(characters, needle)
            if center is not None:
                return center
        return None

    centers: dict[str, float] = {}
    for station, needle in {"DoE": "doe", "BARC": "barc"}.items():
        center = band_center(needle)
        if center is not None:
            centers[station] = center
    darus_center = band_center("darussalam")
    if darus_center is None:
        darus = band_center("darus")
        salam = band_center("salam")
        if darus is not None and salam is not None:
            darus_center = (darus + salam) / 2
    if darus_center is not None:
        centers["Darus-salam"] = darus_center
    if len(centers) < 2:
        return []

    parameter_center = _header_span_center(header[1], "parameter")
    summary_center = _header_span_center(header[1], "summary")
    if parameter_center is None or summary_center is None:
        return []
    parameter_right = (parameter_center + summary_center) / 2

    ordered_centers = sorted(centers.values())
    descriptor_right = ordered_centers[0] - (ordered_centers[1] - ordered_centers[0]) / 4
    station_bounds: dict[str, tuple[float, float]] = {}
    for station, center in centers.items():
        position = ordered_centers.index(center)
        left = descriptor_right if position == 0 else (ordered_centers[position - 1] + center) / 2
        right = (
            center + (center - ordered_centers[position - 1]) / 2
            if position == len(ordered_centers) - 1
            else (center + ordered_centers[position + 1]) / 2
        )
        station_bounds[station] = (left, right)

    parameter_lines: list[tuple[float, str, str]] = []
    for top, characters in lines:
        if top <= header[0]:
            continue
        raw = _chars_text(characters, 0, parameter_right)
        if not re.search(r"[A-Za-z]", raw):
            continue
        nearby_parts: list[str] = []
        for other_top, other_characters in lines:
            part = _chars_text(other_characters, 0, parameter_right)
            if abs(other_top - top) <= 15 and part and part not in nearby_parts:
                nearby_parts.append(part)
        parameter_raw = clean_text(" ".join(nearby_parts))
        pollutant = _pollutant_identity(parameter_raw)
        if pollutant:
            parameter_lines.append((top, pollutant, parameter_raw))
    if not parameter_lines:
        return []

    statistic_lines: list[tuple[float, list[dict[str, Any]], str]] = []
    for top, characters in lines:
        if top <= header[0]:
            continue
        statistic_raw = _chars_text(characters, parameter_right, descriptor_right)
        statistic_compact = _compact(statistic_raw)
        statistic = ""
        if "average" in statistic_compact:
            statistic = "Average"
        elif statistic_compact.startswith("max") or statistic_compact.endswith("max"):
            statistic = "Max"
        elif statistic_compact.startswith("min") or statistic_compact.endswith("min"):
            statistic = "Min"
        elif "excedance" in statistic_compact or "exceedance" in statistic_compact:
            statistic = "Exceedance"
        elif "datacapture" in statistic_compact:
            statistic = "Data capture(%)"
        if not statistic:
            continue
        statistic_lines.append((top, characters, statistic))

    extracted: list[dict[str, Any]] = []
    average_positions = [
        index for index, (_, _, statistic) in enumerate(statistic_lines) if statistic == "Average"
    ]
    for block_number, start in enumerate(average_positions):
        stop = average_positions[block_number + 1] if block_number + 1 < len(average_positions) else len(statistic_lines)
        block = statistic_lines[start:stop]
        block_start = block[0][0]
        block_end = block[-1][0]
        candidates = [
            item for item in parameter_lines if block_start - 2 <= item[0] <= block_end + 2
        ]
        if not candidates:
            continue
        _, pollutant, parameter_raw = candidates[0]
        base_unit = _pollutant_unit(parameter_raw)
        for _, characters, statistic in block:
            statistic_top = min(float(character["top"]) for character in characters)
            value_characters = sorted(
                [
                    character
                    for other_top, other_characters in lines
                    if abs(other_top - statistic_top) <= 6.0
                    for character in other_characters
                ],
                key=lambda item: float(item["x0"]),
            )
            if statistic == "Exceedance":
                statistic = f"Exceedance ({'hours' if pollutant in {'CO', 'O3'} else 'days'})"
            unit = _reported_unit(statistic)
            if unit == "not_stated_in_summary_table":
                unit = base_unit
            for station, (left, right) in station_bounds.items():
                cell_text = re.sub(r"\s+", "", _chars_text(value_characters, left, right))
                dna_match = re.search(r"DNA", cell_text, flags=re.IGNORECASE)
                numeric_matches = re.findall(r"[-+]?\d+(?:\.\d+)?", cell_text)
                raw_value = "DNA" if dna_match else (numeric_matches[-1] if numeric_matches else "")
                if not raw_value:
                    continue
                numeric_value = _reported_value(raw_value)
                extracted.append(
                    {
                        "report_month": source["period"],
                        "station_label_as_reported": station,
                        "parameter": pollutant,
                        "parameter_as_reported": parameter_raw,
                        "statistic_as_reported": statistic,
                        "value_as_reported": raw_value,
                        "value": numeric_value,
                        "unit": unit,
                        "is_missing": numeric_value is None,
                        "page_number": page_number,
                        "table_number": 1,
                        "extraction_method": "pdfplumber_character_positions",
                        "source_url": source["source_url"],
                        "source_sha256": source["sha256"],
                    }
                )
    return extracted


def _extract_monthly_legacy(path: Path, source: dict[str, Any]) -> list[dict[str, Any]]:
    """Extract the fixed 2013-2019 DoE annex schema, including continuation pages."""
    block_schema: list[tuple[str | None, str]] = [
        ("SO2", "SO2 -24 hr"),
        ("NO2", "NO2 -24 hr"),
        ("CO", "CO -1 hr"),
        ("CO", "CO -8 hr"),
        (None, "NOx -1 hr"),
        ("O3", "O3 -8 hr"),
        ("PM2.5", "PM2.5 -24 hr"),
        ("PM10", "PM10 -24 hr"),
    ]
    extracted: list[dict[str, Any]] = []
    block_number = 0
    with pdfplumber.open(path) as document:
        for page_number, page in enumerate(document.pages, start=1):
            for table_number, table in enumerate(page.extract_tables(), start=1):
                average_columns = [
                    index
                    for row in table
                    for index, value in enumerate(row)
                    if _normalized_header(value) == "average"
                ]
                if not average_columns:
                    continue
                summary_index = max(set(average_columns), key=average_columns.count)
                starts = [
                    index
                    for index, row in enumerate(table)
                    if len(row) > summary_index
                    and _normalized_header(row[summary_index]) == "average"
                ]
                if len(starts) < 2:
                    continue
                width = max(len(row) for row in table)
                station_count = width - summary_index - 1
                if station_count == 1:
                    station_columns = [(summary_index + 1, "CAMS-3 (Darus-Salam) AIRQUIS")]
                elif station_count == 2:
                    station_columns = [
                        (summary_index + 1, "CAMS-3 (Darus-Salam) AIRQUIS"),
                        (summary_index + 2, "CAMS-3 (Darus-Salam) Manual"),
                    ]
                else:
                    station_columns = [
                        (summary_index + 1, "CAMS-1 (Sangshad Bhaban)"),
                        (summary_index + 2, "CAMS-2 (BARC/Farmgate)"),
                        (summary_index + 3, "CAMS-3 (Darus-Salam)"),
                    ]
                for local_block, start in enumerate(starts):
                    stop = starts[local_block + 1] if local_block + 1 < len(starts) else len(table)
                    if block_number >= len(block_schema):
                        break
                    visible_parameter = " ".join(
                        clean_text(row[0]) for row in table[start:stop] if row and clean_text(row[0])
                    )
                    visible_pollutant = _pollutant_identity(visible_parameter)
                    if visible_pollutant:
                        matching_index = next(
                            (
                                index
                                for index in range(block_number, len(block_schema))
                                if block_schema[index][0] == visible_pollutant
                            ),
                            block_number,
                        )
                        block_number = matching_index
                    pollutant, parameter_label = block_schema[block_number]
                    block_number += 1
                    if pollutant is None:
                        continue
                    unit_index = max(0, summary_index - 2)
                    source_unit = clean_text(table[start][unit_index])
                    for row in table[start:stop]:
                        padded = list(row) + [None] * max(0, width - len(row))
                        statistic = clean_text(padded[summary_index])
                        if not statistic:
                            continue
                        unit = _reported_unit(statistic)
                        if unit == "not_stated_in_summary_table":
                            unit = source_unit or "not_resolved"
                        for column, station in station_columns:
                            raw_value = clean_text(padded[column])
                            if not raw_value:
                                continue
                            numeric_value = _reported_value(raw_value)
                            extracted.append(
                                {
                                    "report_month": source["period"],
                                    "station_label_as_reported": station,
                                    "parameter": pollutant,
                                    "parameter_as_reported": parameter_label,
                                    "statistic_as_reported": statistic,
                                    "value_as_reported": raw_value,
                                    "value": numeric_value,
                                    "unit": unit,
                                    "is_missing": numeric_value is None,
                                    "page_number": page_number,
                                    "table_number": table_number,
                                    "extraction_method": "pdfplumber_legacy_fixed_schema",
                                    "source_url": source["source_url"],
                                    "source_sha256": source["sha256"],
                                }
                            )
    return extracted


def extract_monthly(path: Path, source: dict[str, Any]) -> list[dict[str, Any]]:
    if path.suffix.lower() != ".pdf":
        raise ValueError("Monthly extractor currently requires a PDF report")
    if int(str(source["period"])[:4]) < 2022:
        return _extract_monthly_legacy(path, source)
    extracted: list[dict[str, Any]] = []
    with pdfplumber.open(path) as document:
        for page_number, page in enumerate(document.pages, start=1):
            page_rows: list[dict[str, Any]] = []
            for table_number, table in enumerate(page.extract_tables(), start=1):
                header_index = None
                for index, row in enumerate(table):
                    normalized = [_normalized_header(cell) for cell in row]
                    if "parameter" in normalized and "summary" in normalized:
                        header_index = index
                        break
                if header_index is None:
                    continue
                header = [clean_text(cell) for cell in table[header_index]]
                normalized_header = [_normalized_header(cell) for cell in header]
                parameter_index = next(
                    (index for index, value in enumerate(normalized_header) if value == "parameter"),
                    None,
                )
                summary_index = next(
                    (index for index, value in enumerate(normalized_header) if value == "summary"),
                    None,
                )
                unit_index = next(
                    (index for index, value in enumerate(normalized_header) if value == "unit"),
                    None,
                )
                if parameter_index is None or summary_index is None:
                    continue
                station_columns = [
                    index
                    for index in range(summary_index + 1, len(header))
                    if _is_dhaka_station(header[index])
                ]
                if not station_columns:
                    continue
                current_parameter = ""
                current_unit = ""
                for row in table[header_index + 1 :]:
                    padded = list(row) + [None] * max(0, len(header) - len(row))
                    if clean_text(padded[parameter_index]):
                        current_parameter = clean_text(padded[parameter_index])
                    if unit_index is not None and clean_text(padded[unit_index]):
                        current_unit = clean_text(padded[unit_index])
                    statistic = clean_text(padded[summary_index])
                    if not current_parameter or not statistic:
                        continue
                    pollutant = _pollutant_identity(current_parameter)
                    if pollutant is None:
                        continue
                    for column in station_columns:
                        raw_value = clean_text(padded[column])
                        unit = _reported_unit(statistic)
                        if unit == "not_stated_in_summary_table":
                            unit = current_unit or _pollutant_unit(current_parameter)
                        page_rows.append(
                            {
                                "report_month": source["period"],
                                "station_label_as_reported": header[column],
                                "parameter": pollutant,
                                "parameter_as_reported": current_parameter,
                                "statistic_as_reported": statistic,
                                "value_as_reported": raw_value,
                                "value": _reported_value(raw_value),
                                "unit": unit,
                                "is_missing": _reported_value(raw_value) is None,
                                "page_number": page_number,
                                "table_number": table_number,
                                "extraction_method": "pdfplumber_table",
                                "source_url": source["source_url"],
                                "source_sha256": source["sha256"],
                            }
                        )
            extracted.extend(page_rows or _extract_monthly_positional(page, page_number, source))
    unique: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in extracted:
        key = (
            row["report_month"],
            row["station_label_as_reported"],
            row["parameter_as_reported"],
            row["statistic_as_reported"],
            row["source_url"],
        )
        unique[key] = row
    return list(unique.values())


def extract_monthly_report_aqi(path: Path, source: dict[str, Any]) -> list[dict[str, Any]]:
    """Extract Dhaka's daily numeric AQI from Table 6 in newer monthly PDFs."""
    if path.suffix.lower() != ".pdf" or int(str(source["period"])[:4]) < 2022:
        return []
    try:
        result = subprocess.run(
            ["pdftotext", "-layout", str(path), "-"],
            check=True,
            capture_output=True,
            text=True,
        )
    except FileNotFoundError as error:
        raise RuntimeError("Poppler pdftotext is required for rotated AQI tables") from error
    except subprocess.CalledProcessError as error:
        raise RuntimeError(f"pdftotext failed for {path}: {error.stderr.strip()}") from error

    report_year, report_month = (int(part) for part in str(source["period"]).split("-"))
    table_match = re.search(
        r"(?im)^\s*Table\s*6\s*:?\s*Air\s+Quality\s+Index\s*\(AQI\)[^\n]*$",
        result.stdout,
    )
    if table_match is None:
        return []
    table_end = re.search(r"(?im)^\s*(?:Table\s*7\b|Figure\s+\d+\b)", result.stdout[table_match.end() :])
    end = table_match.end() + table_end.start() if table_end else len(result.stdout)
    table_text = result.stdout[table_match.end() : end]
    row_pattern = re.compile(
        r"^\s*(\d{1,2})\s*[-/.]\s*(\d{1,2})\s*[-/.]\s*(\d{2,4})\s+"
        r"(DNA|N/?A|\d{1,3})(?:\s|$)",
        flags=re.IGNORECASE | re.MULTILINE,
    )
    rows: dict[str, dict[str, Any]] = {}
    for match in row_pattern.finditer(table_text):
        first, second, year_text = (int(match.group(index)) for index in range(1, 4))
        year = year_text + 2000 if year_text < 100 else year_text
        if year != report_year:
            continue
        if first == report_month:
            day = second
        elif second == report_month:
            day = first
        else:
            continue
        try:
            observation_date = datetime(year, report_month, day).date().isoformat()
        except ValueError:
            continue
        raw_aqi = match.group(4).upper()
        aqi = int(raw_aqi) if raw_aqi.isdigit() else None
        rows[observation_date] = {
            "report_month": source["period"],
            "aqi_date": observation_date,
            "city": "Dhaka",
            "aqi_as_reported": raw_aqi,
            "aqi": aqi,
            "is_missing": aqi is None,
            "extraction_method": "poppler_pdftotext_layout_table_6",
            "source_url": source["source_url"],
            "source_sha256": source["sha256"],
        }
    return list(rows.values())


def _season(month: int) -> str:
    if month in {12, 1, 2}:
        return "Winter"
    if month in {3, 4, 5}:
        return "Pre-monsoon"
    if month in {6, 7, 8, 9}:
        return "Monsoon"
    return "Post-monsoon"


def _effective_measurement_unit(row: pd.Series) -> str:
    parameter_unit = _pollutant_unit(str(row["parameter_as_reported"]))
    if parameter_unit != "not_resolved":
        return parameter_unit
    source_unit = clean_text(row["unit"])
    if source_unit in {"µg /m3", "µg/m3", "ug/m3"}:
        return "ug/m3"
    return source_unit


def build_monthly_dataset(
    monthly: pd.DataFrame,
    daily: pd.DataFrame,
    monthly_report_aqi: pd.DataFrame,
) -> pd.DataFrame:
    """Build an analysis-friendly monthly table shaped like the original dataset."""
    monthly_periods = pd.to_datetime(monthly["report_month"], format="%Y-%m")
    daily_periods = pd.to_datetime(daily["report_date"]).dt.to_period("M").dt.to_timestamp()
    start = monthly_periods.min()
    end = max(monthly_periods.max(), daily_periods.max())
    month_index = pd.date_range(start, end, freq="MS")
    wide = pd.DataFrame({"month_start": month_index.strftime("%Y-%m-%d")})
    wide["year"] = month_index.year
    wide["month"] = month_index.month
    wide["month_name"] = month_index.month_name()
    wide["season"] = [_season(month) for month in month_index.month]

    pollutant_specs = {
        "pm25": ("PM2.5", "ug/m3"),
        "pm10": ("PM10", "ug/m3"),
        "no2": ("NO2", "ppb"),
        "so2": ("SO2", "ppb"),
        "co": ("CO", "ppm"),
        "o3": ("O3", "ppb"),
    }
    source = monthly.copy()
    source["effective_unit"] = source.apply(_effective_measurement_unit, axis=1)
    source["statistic"] = source["statistic_as_reported"].map(_normalized_header)
    for prefix, (parameter, expected_unit) in pollutant_specs.items():
        subset = source[source["parameter"].eq(parameter)].copy()
        if parameter == "CO":
            subset = subset[
                subset["parameter_as_reported"].map(_compact).str.contains("co8hr", na=False)
            ]
        grouped_rows: list[dict[str, Any]] = []
        for period, group in subset.groupby("report_month"):
            measurements = group[
                group["effective_unit"].isin({expected_unit, "not_resolved"})
            ]
            averages = measurements[measurements["statistic"].eq("average")]["value"].dropna()
            minima = measurements[measurements["statistic"].eq("min")]["value"].dropna()
            maxima = measurements[measurements["statistic"].eq("max")]["value"].dropna()
            capture = group[group["statistic"].str.contains("data capture", na=False)][
                "value"
            ].dropna()
            units = sorted(
                set(
                    measurements.loc[
                        measurements["statistic"].isin({"average", "min", "max"}),
                        "effective_unit",
                    ]
                )
            )
            grouped_rows.append(
                {
                    "month_start": f"{period}-01",
                    f"{prefix}_mean": averages.mean() if not averages.empty else None,
                    f"{prefix}_median": None,
                    f"{prefix}_min": minima.min() if not minima.empty else None,
                    f"{prefix}_max": maxima.max() if not maxima.empty else None,
                    f"{prefix}_station_count": int(averages.count()),
                    f"{prefix}_mean_data_capture_pct": capture.mean() if not capture.empty else None,
                    f"{prefix}_unit_as_reported": "; ".join(units),
                    f"{prefix}_unit_status": (
                        "resolved"
                        if units == [expected_unit]
                        else "partly_or_fully_unresolved_in_summary_table"
                    ),
                }
            )
        grouped = pd.DataFrame(grouped_rows)
        if not grouped.empty:
            wide = wide.merge(grouped, on="month_start", how="left")
        else:
            for suffix in [
                "mean", "median", "min", "max", "station_count",
                "mean_data_capture_pct", "unit_as_reported", "unit_status",
            ]:
                wide[f"{prefix}_{suffix}"] = None

    report_aqi = monthly_report_aqi.copy()
    if not report_aqi.empty:
        report_aqi = report_aqi[report_aqi["aqi"].notna()]
        report_aqi["month_start"] = report_aqi["report_month"].astype(str) + "-01"
    selected_daily = daily[daily["selected_record"].astype(str).str.lower().isin({"true", "1"})].copy()
    selected_daily = selected_daily[selected_daily["aqi"].notna()]
    selected_daily["month_start"] = pd.to_datetime(selected_daily["report_date"]).dt.strftime(
        "%Y-%m-01"
    )

    aqi_rows: list[dict[str, Any]] = []
    for month_start in wide["month_start"]:
        candidates: list[tuple[str, pd.Series]] = []
        for basis, frame in [
            ("monthly_report_table_6", report_aqi),
            ("standalone_daily_archive_report_date", selected_daily),
        ]:
            values = frame.loc[frame["month_start"].eq(month_start), "aqi"].dropna()
            if not values.empty:
                candidates.append((basis, values))
        if not candidates:
            aqi_rows.append({"month_start": month_start})
            continue
        basis, values = max(candidates, key=lambda item: (len(item[1]), item[0].startswith("monthly")))
        timestamp = pd.Timestamp(month_start)
        calendar_days = calendar.monthrange(timestamp.year, timestamp.month)[1]
        aqi_rows.append(
            {
                "month_start": month_start,
                "aqi_mean": values.mean(),
                "aqi_median": values.median(),
                "aqi_min": values.min(),
                "aqi_max": values.max(),
                "aqi_days_reported": int(values.count()),
                "aqi_calendar_days": calendar_days,
                "aqi_coverage_pct": 100 * values.count() / calendar_days,
                "aqi_source_basis": basis,
            }
        )
    wide = wide.merge(pd.DataFrame(aqi_rows), on="month_start", how="left")
    wide["pollutant_aggregation_basis"] = (
        "mean=unweighted mean of station monthly averages; min/max=extreme station monthly "
        "values; pollutant medians unavailable in source reports"
    )
    wide["geography"] = "Dhaka reporting stations/city AQI"
    core_columns = ["month_start", "year", "month", "month_name", "season"]
    for prefix in pollutant_specs:
        core_columns.extend(
            [f"{prefix}_mean", f"{prefix}_median", f"{prefix}_min", f"{prefix}_max"]
        )
    core_columns.extend(["aqi_mean", "aqi_median", "aqi_min", "aqi_max"])
    return wide[core_columns + [column for column in wide if column not in core_columns]]


def select_daily_records(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    frame = pd.DataFrame(rows)
    if frame.empty:
        return rows
    for _, indices in frame.groupby("report_date", sort=False).groups.items():
        group_indices = list(indices)
        duplicate = len(group_indices) > 1
        for index in group_indices:
            rows[index]["duplicate_date"] = duplicate
        signatures = {
            (
                rows[index]["aqi"],
                rows[index]["responsible_pollutant"],
                rows[index]["aqi_category_as_reported"],
            )
            for index in group_indices
        }
        if len(signatures) == 1:
            rows[group_indices[0]]["selected_record"] = True
            if duplicate:
                for index in group_indices:
                    rows[index]["qa_status"] = "duplicate_date_same_values"
        else:
            for index in group_indices:
                rows[index]["qa_status"] = "conflicting_duplicate_date"
    return rows


def write_workbook(
    output: Path,
    monthly_dataset: pd.DataFrame,
    daily: pd.DataFrame,
    monthly_report_aqi: pd.DataFrame,
    monthly: pd.DataFrame,
    population: pd.DataFrame,
    population_worldometer: pd.DataFrame,
    tree_cover_loss: pd.DataFrame,
    hdi: pd.DataFrame,
    manifest: pd.DataFrame,
    qa: pd.DataFrame,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    readme = workbook.create_sheet("read_me")
    readme_rows = [
        ("Workbook", "Dhaka DoE air-quality data extracted from official public reports"),
        ("Daily source", DAILY_ARCHIVE_URL),
        ("Monthly source", MONTHLY_ARCHIVE_URL),
        ("Daily meaning", "Published Dhaka city AQI; not a pollutant concentration"),
        ("Daily report_date", "Date listed in the DoE archive; compare document_aqi_date before time-series use"),
        ("AQI pollutant", "responsible_pollutant is the pollutant controlling the published AQI, not a full driver decomposition"),
        ("Main sheet", "monthly_dataset follows the original wide monthly structure using only official DoE values"),
        ("Monthly meaning", "Station-level reported summary statistics; units are not inferred when absent"),
        ("Pollutant aggregation", "Monthly means average the reporting-station monthly averages; minima/maxima are station extremes"),
        ("Pollutant medians", "Retained for schema compatibility but blank because the source reports do not publish them"),
        ("AQI coverage", "Numeric monthly AQI uses DoE Table 6 or the standalone daily archive; earlier unavailable values remain blank"),
        ("Monthly coverage", "Official linked reports: 2013-2019 and 2022 onward; no 2020-2021 year pages are linked"),
        ("Monthly pollutants", "PM2.5, PM10, SO2, NO2, CO, and O3 are retained as separate reported measurements"),
        (
            "National context",
            "Separate national Bangladesh sheets. Population is the complete official UN WUP 2025 "
            "series; rural equals total minus urban and is checked against the reported rural series. "
            "Worldometer is retained only as a sparse UN-derived cross-check. Global Forest Watch "
            "tree-cover loss and HDI are descriptive context and are not used as causal predictors.",
        ),
        ("Missing marker", "DoE uses DNA; numeric value is blank and is_missing is TRUE"),
        ("Duplicate policy", "All attachments retained; selected_record is TRUE only when a date is unambiguous"),
        ("Category schemes", "DoE category labels changed over time; source_category_scheme and source labels are preserved"),
        ("QA", "Use qa_issues and extraction_status before analysis; partial means at least one pollutant block was not parsed"),
        ("Important", "Do not combine daily AQI and monthly concentrations as one measurement series"),
    ]
    for row in readme_rows:
        readme.append(row)
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 110
    readme.freeze_panes = "A2"
    for cell in readme[1]:
        cell.font = Font(bold=True)

    _write_frame(workbook, "monthly_dataset", monthly_dataset)
    _write_frame(workbook, "daily_dhaka_aqi", daily)
    _write_frame(workbook, "monthly_report_aqi", monthly_report_aqi)
    _write_frame(workbook, "monthly_dhaka", monthly)
    _write_frame(workbook, "population", population)
    _write_frame(workbook, "population_worldometer", population_worldometer)
    _write_frame(workbook, "tree_cover_loss", tree_cover_loss)
    _write_frame(workbook, "hdi", hdi)
    _write_frame(workbook, "source_manifest", manifest)
    _write_frame(workbook, "qa_issues", qa)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def _write_frame(workbook: Workbook, name: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(frame.columns.tolist())
    for row in frame.itertuples(index=False, name=None):
        sheet.append([None if pd.isna(value) else value for value in row])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1 and sheet.max_column > 0:
        sheet.auto_filter.ref = sheet.dimensions
    for column_index, column_name in enumerate(frame.columns, start=1):
        samples = [str(column_name)] + [
            str(sheet.cell(row=row, column=column_index).value or "")
            for row in range(2, min(sheet.max_row, 102) + 1)
        ]
        width = min(60, max(10, max(len(value) for value in samples) + 2))
        sheet.column_dimensions[get_column_letter(column_index)].width = width
        if "url" in str(column_name).lower():
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=column_index)
                if isinstance(cell.value, str) and cell.value.startswith("https://"):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
