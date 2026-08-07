import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.doe import discover_monthly_year_pages

ROOT = Path(__file__).resolve().parents[1]
PROCESSED = ROOT / "data/processed"
POLLUTANTS = {"PM2.5", "PM10", "SO2", "NO2", "CO", "O3"}


def test_markdown_is_limited_to_project_readme_and_manuscript() -> None:
    markdown_files = {
        path.relative_to(ROOT)
        for path in ROOT.rglob("*.md")
        if not {".git", ".venv", ".pytest_cache"}.intersection(path.parts)
    }
    assert markdown_files == {
        Path("README.md"),
        Path("paper/draft/manuscript.md"),
    }

    readme = (ROOT / "README.md").read_text(encoding="utf-8")
    manuscript = (ROOT / "paper/draft/manuscript.md").read_text(encoding="utf-8")
    assert r"$\text{PM}_{2.5}$" in readme or r"$\mathrm{PM}_{2.5}$" in readme
    assert r"$\text{PM}_{10}$" in readme or r"$\mathrm{PM}_{10}$" in readme
    assert "PM₂.₅" in manuscript
    assert "PM₁₀" in manuscript
    assert r"$\mathrm{PM}" not in manuscript
    assert len(re.findall(r"(?<!\$)\$(?!\$).+?(?<!\$)\$(?!\$)", manuscript)) <= 5
    assert "PM~" not in readme
    assert "PM~" not in manuscript
    assert r"\hat{\beta}_1" in manuscript
    assert r"E_{150}=\sum" in manuscript
    assert "**Awnon Bhowmik**" in manuscript
    assert (
        "*Department of Computer Science and Engineering, "
        "Colorado Technical University*"
    ) in manuscript
    assert "awnonbhowmik@outlook.com" in manuscript

    references = manuscript.split("## References", maxsplit=1)[1]
    assert len(re.findall(r"^\d+\. ", references, flags=re.MULTILINE)) >= 30
    for scholarly_doi in (
        "10.1073/pnas.1803222115",
        "10.1016/j.scitotenv.2018.07.288",
        "10.1513/AnnalsATS.202103-252OC",
        "10.4209/aaqr.220082",
        "10.1016/j.atmosenv.2023.119587",
    ):
        assert scholarly_doi in references


def test_monthly_master_page_discovery() -> None:
    html = """
    <a href="/pages/static-pages/monthly-air-quality-report-2026-id">
      Monthly Air Quality Report 2026
    </a>
    <a href="https://example.com/not-doe">Monthly Air Quality Report 2025</a>
    <a href="/pages/other/not-static">Monthly Air Quality Report 2024</a>
    """
    assert discover_monthly_year_pages(html) == {
        2026: "https://doe.gov.bd/pages/static-pages/monthly-air-quality-report-2026-id"
    }


def test_daily_archive_selection_and_lineage() -> None:
    daily = pd.read_csv(PROCESSED / "doe_daily_dhaka_aqi.csv")
    assert len(daily) > 1_100
    assert daily["aqi"].notna().all()
    assert daily["source_sha256"].str.fullmatch(r"[0-9a-f]{64}").all()
    assert daily["source_url"].str.startswith("https://").all()
    for _, group in daily.groupby("report_date"):
        if group["qa_status"].eq("conflicting_duplicate_date").any():
            assert not group["selected_record"].any()
        else:
            assert group["selected_record"].sum() == 1


def test_monthly_pollutants_and_units_are_separate() -> None:
    monthly = pd.read_csv(PROCESSED / "doe_monthly_dhaka.csv")
    assert monthly["report_month"].min() == "2013-01"
    assert POLLUTANTS == set(monthly["parameter"])
    assert monthly["source_sha256"].str.fullmatch(r"[0-9a-f]{64}").all()
    assert set(monthly["unit"]) - {"days", "hours", "percent"}
    assert not monthly["parameter"].str.contains("AQI", case=False).any()


def test_manifest_and_workbook_are_consistent() -> None:
    manifest = pd.read_csv(PROCESSED / "doe_source_manifest.csv")
    assert set(manifest["source_kind"]) == {"daily_aqi", "monthly_report"}
    assert manifest["sha256"].str.fullmatch(r"[0-9a-f]{64}").all()
    assert not manifest["extraction_status"].eq("failed").any()

    workbook = load_workbook(PROCESSED / "dhaka_doe_air_quality.xlsx", read_only=True)
    assert workbook.sheetnames == [
        "read_me",
        "monthly_dataset",
        "daily_dhaka_aqi",
        "monthly_report_aqi",
        "monthly_dhaka",
        "population",
        "population_worldometer",
        "tree_cover_loss",
        "hdi",
        "source_manifest",
        "qa_issues",
    ]
    assert (
        workbook["daily_dhaka_aqi"].max_row
        == len(pd.read_csv(PROCESSED / "doe_daily_dhaka_aqi.csv")) + 1
    )
    assert (
        workbook["monthly_dhaka"].max_row
        == len(pd.read_csv(PROCESSED / "doe_monthly_dhaka.csv")) + 1
    )


def test_wide_monthly_dataset_preserves_original_shape_without_fabrication() -> None:
    wide = pd.read_csv(PROCESSED / "doe_monthly_dataset_wide.csv")
    expected = {
        "month_start",
        "year",
        "month",
        "season",
        "pm25_mean",
        "pm25_median",
        "pm25_min",
        "pm25_max",
        "pm10_mean",
        "no2_mean",
        "so2_mean",
        "co_mean",
        "o3_mean",
        "aqi_mean",
        "aqi_median",
        "aqi_min",
        "aqi_max",
    }
    assert expected.issubset(wide.columns)
    assert wide["month_start"].min() == "2013-01-01"
    assert wide["month_start"].is_unique
    assert wide.loc[wide["year"].lt(2022), "aqi_mean"].isna().all()
    assert wide.loc[wide["year"].ge(2022), "aqi_mean"].notna().all()
    for pollutant in ["pm25", "pm10", "no2", "so2", "co", "o3"]:
        assert wide[f"{pollutant}_mean"].notna().any()
        assert wide[f"{pollutant}_median"].isna().all()
    assert wide["aqi_coverage_pct"].dropna().between(0, 100).all()


def test_population_and_hdi_provenance_and_year_alignment() -> None:
    population = pd.read_csv(ROOT / "data/context/bangladesh_population.csv")
    assert population["year"].tolist() == list(range(2013, 2026))
    assert population["geographic_scope"].eq("national").all()
    assert (
        population["total_population"]
        .sub(population["urban_population"])
        .eq(population["rural_population"])
        .all()
    )
    assert population["rural_population"].eq(population["rural_population_un_reported"]).all()
    assert (
        population["urban_population"]
        .div(population["total_population"])
        .sub(population["urban_share_fraction"])
        .abs()
        .lt(1e-8)
        .all()
    )
    assert population["source_url"].str.startswith("https://population.un.org/").all()
    assert population["source_sha256"].str.fullmatch(r"[0-9a-f]{64}").all()

    context = pd.read_csv(ROOT / "data/context/bangladesh_hdi.csv")
    assert context["retained_source_url"].str.startswith("https://").all()
    assert context["undp_verification_url"].str.startswith("https://").all()
    assert context["year"].min() == 2013
    assert context.loc[context["year"].le(2023), "hdi_undp_same_year"].notna().all()
    assert context.loc[context["year"].ge(2024), "hdi_undp_same_year"].isna().all()
    assert (
        context.loc[context["year"].eq(2025), "hdi_observation_year_for_retained_value"].item()
        == 2023
    )
    assert (
        context.loc[context["year"].between(2017, 2024), "retained_source"]
        .eq("AIDS_BD_2000_2024.xlsx")
        .all()
    )
    assert (
        context.loc[context["year"].between(2023, 2024), "hdi_observation_year_for_retained_value"]
        .isna()
        .all()
    )

    workbook = load_workbook(PROCESSED / "dhaka_doe_air_quality.xlsx", read_only=True)
    population_headers = [cell.value for cell in workbook["population"][1]]
    hdi_headers = [cell.value for cell in workbook["hdi"][1]]
    assert "rural_population" in population_headers
    assert "hdi_undp_same_year" in hdi_headers


def test_worldometer_and_tree_cover_context_are_scoped_and_traceable() -> None:
    worldometer = pd.read_csv(ROOT / "data/context/bangladesh_population_worldometer.csv")
    assert worldometer["year"].tolist() == [2010, 2015, 2020, 2022, 2023, 2024, 2025, 2026]
    assert worldometer["geographic_scope"].eq("national").all()
    assert worldometer["analysis_role"].eq("descriptive_cross_check").all()
    assert worldometer["source_url"].str.contains("worldometers.info").all()

    forest = pd.read_csv(ROOT / "data/context/bangladesh_tree_cover_loss.csv")
    assert forest["year"].tolist() == list(range(2001, 2025))
    assert forest["geographic_scope"].eq("national").all()
    assert forest["metric"].eq("tree_cover_loss_all_causes").all()
    assert forest["provider"].eq("Global Forest Watch").all()
    assert forest["source_sha256"].str.fullmatch(r"[0-9a-f]{64}").all()
    assert forest["definition_note"].str.contains("not synonymous").all()

    workbook = load_workbook(ROOT / "data/processed/dhaka_doe_air_quality.xlsx")
    assert "population_worldometer" in workbook.sheetnames
    assert "tree_cover_loss" in workbook.sheetnames


def test_monthly_report_aqi_lineage() -> None:
    aqi = pd.read_csv(PROCESSED / "doe_monthly_report_dhaka_aqi.csv")
    assert aqi["report_month"].min() == "2022-01"
    assert aqi["aqi_date"].min() == "2022-01-01"
    assert aqi["source_sha256"].str.fullmatch(r"[0-9a-f]{64}").all()
    assert aqi["extraction_method"].eq("poppler_pdftotext_layout_table_6").all()


def test_integrated_analysis_outputs_are_complete() -> None:
    analysis_source = (ROOT / "scripts/analyze_doe_dataset.py").read_text(encoding="utf-8")
    for number in range(1, 9):
        assert f'"Figure {number}.' not in analysis_source

    figure_names = sorted(path.name for path in (ROOT / "analysis/figures").glob("*.png"))
    assert figure_names == [
        "figure_1_coverage_and_monitoring.png",
        "figure_2_daily_aqi_burden.png",
        "figure_3_seasonal_particulate_burden.png",
        "figure_4_pollution_episodes.png",
        "figure_5_trend_sensitivity.png",
        "figure_6_station_heterogeneity.png",
        "figure_7_multipollutant_structure.png",
        "figure_8_contextual_framework.png",
    ]

    workbook = load_workbook(ROOT / "analysis/dhaka_doe_analysis.xlsx", read_only=True)
    expected = {
        "episodes_gt_150",
        "trend_sensitivity",
        "station_sensitivity",
        "particulate_era_summary",
        "particulate_annual",
        "particulate_era_season",
        "daily_unified",
        "daily_source_concordance",
        "station_summary",
        "station_paired_contrasts",
        "historical_pollutants",
        "historical_correlations",
        "recent_correlations",
        "context_change_summary",
        "population_source_comparison",
        "population_worldometer",
        "tree_cover_loss",
    }
    assert expected.issubset(workbook.sheetnames)
    assert "forecast_backtest" not in workbook.sheetnames
    assert "context_assoc" not in workbook.sheetnames

    daily = pd.read_excel(ROOT / "analysis/dhaka_doe_analysis.xlsx", sheet_name="daily_unified")
    assert daily["report_date"].min() == pd.Timestamp("2022-01-01")
    assert daily["report_date"].max() == pd.Timestamp("2025-12-31")
    assert len(daily) == 1457
    assert daily["report_date"].is_unique

    episodes = pd.read_excel(
        ROOT / "analysis/dhaka_doe_analysis.xlsx", sheet_name="episodes_gt_150"
    )
    longest = episodes.nlargest(1, "duration_days").iloc[0]
    assert longest["duration_days"] == 145
    assert longest["start_date"] == pd.Timestamp("2022-10-26")
    assert longest["end_date"] == pd.Timestamp("2023-03-19")
